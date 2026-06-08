from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from ui.ui_page_arxiki import Ui_page
from PySide2.QtWidgets import QWidget, QMessageBox, QAbstractItemView, QFileDialog
from PySide2.QtCore import Qt, QSortFilterProxyModel, QRegExp, QTimer
from PySide2.QtGui import QStandardItemModel, QStandardItem, QColor, QRegExpValidator, QFont
from delegates.delegates_dlt import DeleteButtonDelegate, EditButtonDelegate
from utils.message import show_temp_message

from database_manager import delete_producer

class arxikiPage(QWidget):
    def __init__(self, parent=None, main_window=None):
        super().__init__(parent)

        # Φόρτωμα UI από Designer
        self.ui=Ui_page()
       
        self.ui.setupUi(self)
        self.model = QStandardItemModel(self)
        self.ui.tableView.setModel(self.model)

        self.model.setHorizontalHeaderLabels([
                                              "ΑΦΜ",
                                              "Όνομα",
                                              "Επώνυμο",
                                              "Περιφέρεια",
                                              "Αρχική ΤΑ",
                                              "Μελλοντική ΤΑ",
                                              "Μόρια",
                                              "Επιλεξιμότητα",
                                              "Τελευταία\nΕπεξεργασία",
                                              "Επ/σία",
                                              "Διαγραφή",
                                                ])
        
        
        #Για την προσθήκη φίλτρου
        self.proxy = QSortFilterProxyModel(self)
        self.proxy.setSourceModel(self.model)

        #  φιλτράρουμε στη στήλη ΑΦΜ (στήλη 0)
        self.proxy.setFilterKeyColumn(0)
        self.proxy.setFilterCaseSensitivity(Qt.CaseInsensitive)

        self.ui.tableView.setModel(self.proxy)
        self.ui.lineEdit_search.textChanged.connect(self.filter_afm)

        #περιορισμός έως 9 ψηφίων στο lineEdit του ΑΦΜ search
        regex = QRegExp(r"\d{0,9}")  # από 0 έως 9 ψηφία
        validator = QRegExpValidator(regex)
        self.ui.lineEdit_search.setValidator(validator)
        
        #Kουμπί διαγραφής πίνακα
        self.delete_delegate = DeleteButtonDelegate(self.ui.tableView)
        self.delete_delegate.deleteClicked.connect(self.delete_btn)
        self.ui.tableView.setItemDelegateForColumn(10, self.delete_delegate)

        #Κουμπί επεξεργασίας πίνακα
        self.edit_delegate = EditButtonDelegate(self.ui.tableView)
        self.edit_delegate.editClicked.connect(self.edit_btn)
        self.ui.tableView.setItemDelegateForColumn(9, self.edit_delegate)

        #Κουμπί εκτύπωσης πίνακα
        self.ui.pushButton.clicked.connect(self.export_table_to_xlsx)

        #σύνδεση με main-window
        self.main_window=main_window

        # Deferred load: πρώτα δείχνουμε το παράθυρο, μετά γεμίζουμε τον πίνακα
        # — έτσι το startup δεν καθυστερεί με πολλές εγγραφές.
        QTimer.singleShot(0, self.load_producers)

        #Table compatibility settings
        self._setup_table_windows_compatibility_arxiki()

        

    _ROW_BG = QColor("#E3F2FD")  # cached: αποφεύγουμε QColor() ανά κελί
    _BOLD_FONT = QFont()
    _BOLD_FONT.setBold(True)

    @staticmethod
    def _format_timestamp(iso_str):
        """ISO 'YYYY-MM-DD HH:MM:SS' → Greek UI 'DD/MM/YYYY HH:MM'. Κενό αν None/άδειο."""
        if not iso_str:
            return ""
        try:
            dt = datetime.strptime(iso_str, "%Y-%m-%d %H:%M:%S")
            return dt.strftime("%d/%m/%Y %H:%M")
        except (ValueError, TypeError):
            return ""

    def load_producers(self, progress_callback=None):
        """Φέρνει όλους τους παραγωγούς από τη βάση στον πίνακα της αρχικής.
        Optional progress_callback(current, total) για ένδειξη προόδου κατά το φόρτωμα."""
        try:
            from database_manager import fetch_all_producers

            tableView = self.ui.tableView
            bg = self._ROW_BG  # τοπική αναφορά — γρηγορότερο από attr lookup στο loop

            tableView.setUpdatesEnabled(False)

            # Πλήρες model reset ΜΕ signals: το QStandardItemModel.clear() καλεί
            # internally beginResetModel/endResetModel, οπότε ο proxy απελευθερώνει
            # persistent indexes και filter mapping — αποφυγή dangling pointers.
            header_labels = [
                self.model.horizontalHeaderItem(i).text() if self.model.horizontalHeaderItem(i) else ""
                for i in range(self.model.columnCount())
            ]
            self.model.clear()
            self.model.setHorizontalHeaderLabels(header_labels)

            # Bulk append με blocked signals (non-destructive — δεν αφήνει dangling
            # pointers γιατί δεν διαγράφουμε τίποτα, μόνο προσθέτουμε).
            self.model.blockSignals(True)
            try:
                producers = fetch_all_producers()
                total = len(producers)
                BATCH = 200  # μικρότερο batch → πιο ομαλή μπάρα progress

                for start in range(0, total, BATCH):
                    chunk = producers[start:start + BATCH]
                    for p in chunk:
                        # p[0]=ΑΦΜ, p[1]=Όνομα, p[2]=Επώνυμο, p[3]=Μόρια, p[4]=Περιφέρεια,
                        # p[5]=Αρχική ΤΑ, p[6]=Μελλοντική ΤΑ, p[7]=Επιλεξιμότητα (combined),
                        # p[8]=last_modified (ISO timestamp)
                        elig_text = str(p[7]) if p[7] else ""
                        row_items = [
                            QStandardItem(str(p[0])),
                            QStandardItem(str(p[1])),
                            QStandardItem(str(p[2])),
                            QStandardItem(str(p[4]) if p[4] else ""),  # Περιφέρεια
                            QStandardItem("" if p[5] is None else str(p[5])),  # Αρχική ΤΑ — fix
                            QStandardItem("" if p[6] is None else str(p[6])),  # Μελλοντική ΤΑ — fix
                            QStandardItem("" if p[3] is None else str(p[3])),  # Μόρια — fix
                            QStandardItem(elig_text),                  # Επιλεξιμότητα
                            QStandardItem(self._format_timestamp(p[8])),  # Τελευταία Επεξεργασία
                            QStandardItem(""),  # Επ/σία
                            QStandardItem("")   # Διαγραφή
                        ]

                        for item in row_items[:9]:
                            item.setEditable(False)
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setBackground(bg)

                        if elig_text == "ΕΠΙΛΕΞΙΜΟΣ":
                            row_items[7].setForeground(QColor("#2E7D32"))
                            row_items[7].setFont(self._BOLD_FONT)
                        elif elig_text == "ΜΗ ΕΠΙΛΕΞΙΜΟΣ":
                            row_items[7].setForeground(QColor("#C62828"))
                            row_items[7].setFont(self._BOLD_FONT)

                        self.model.appendRow(row_items)

                    if progress_callback:
                        progress_callback(min(start + BATCH, total), total)
            finally:
                self.model.blockSignals(False)
                # Αναγκαστικός rebuild του proxy mapping βασισμένος στο νέο source state.
                self.proxy.invalidate()
                # Επαναφορά column widths — το model.clear() τα μηδενίζει μέσω modelReset.
                tableView.setColumnWidth(1, 120)
                tableView.setColumnWidth(2, 120)
                tableView.setColumnWidth(5, 130)
                tableView.setColumnWidth(7, 150)
                tableView.setColumnWidth(8, 160)
                tableView.setUpdatesEnabled(True)
        except Exception as e:
            QMessageBox.critical(
                self,
                "Σφάλμα Φόρτωσης",
                f"Αποτυχία φόρτωσης λίστας παραγωγών.\n\n"
                f"{str(e)}"
            )
                
        
    
    def upsert_producer_row(self, afm):
        """Ανανεώνει ή προσθέτει μία μόνο γραμμή για τον δοσμένο ΑΦΜ, αποφεύγοντας πλήρες reload."""
        try:
            from database_manager import fetch_single_producer_row

            p = fetch_single_producer_row(afm)
            if not p:
                return

            elig_text = str(p[7]) if p[7] else ""
            values = [
                str(p[0]),                        # col 0: ΑΦΜ
                str(p[1]),                        # col 1: Όνομα
                str(p[2]),                        # col 2: Επώνυμο
                str(p[4]) if p[4] else "",        # col 3: Περιφέρεια
                "" if p[5] is None else str(p[5]),          # Αρχική ΤΑ — fix
                "" if p[6] is None else str(p[6]),          # Μελλοντική ΤΑ — fix
                "" if p[3] is None else str(p[3]),          # Μόρια — fix
                elig_text,                        # col 7: Επιλεξιμότητα
                self._format_timestamp(p[8]),     # col 8: Τελευταία Επεξεργασία
            ]

            if elig_text == "ΕΠΙΛΕΞΙΜΟΣ":
                elig_color = QColor("#2E7D32")
            elif elig_text == "ΜΗ ΕΠΙΛΕΞΙΜΟΣ":
                elig_color = QColor("#C62828")
            else:
                elig_color = None

            matches = self.model.findItems(str(p[0]), Qt.MatchExactly, 0)
            if matches:
                row = matches[0].row()
                for col, val in enumerate(values):
                    item = self.model.item(row, col)
                    if item is not None:
                        item.setText(val)
                elig_item = self.model.item(row, 7)
                if elig_item is not None:
                    elig_item.setForeground(elig_color if elig_color else QColor("#000000"))
                    elig_item.setFont(self._BOLD_FONT if elig_color else QFont())
                return

            row_items = [QStandardItem(v) for v in values] + [
                QStandardItem(""), QStandardItem("")   # Επ/σία (col 9), Διαγραφή (col 10)
            ]
            for item in row_items[:9]:
                item.setEditable(False)
                item.setTextAlignment(Qt.AlignCenter)
                item.setBackground(QColor("#E3F2FD"))
            if elig_color is not None:
                row_items[7].setForeground(elig_color)
                row_items[7].setFont(self._BOLD_FONT)

            # Sorted insert: όλα τα ΑΦΜ είναι 9-ψήφια, άρα lexicographic compare
            # ισοδυναμεί με numeric. Διατηρεί την αύξουσα σειρά του fetch_all_producers.
            new_afm = str(p[0])
            insert_at = self.model.rowCount()
            for r in range(self.model.rowCount()):
                existing = self.model.item(r, 0)
                if existing is not None and existing.text() > new_afm:
                    insert_at = r
                    break
            self.model.insertRow(insert_at, row_items)
        except Exception as e:
            QMessageBox.critical(
                self,
                "Σφάλμα Ενημέρωσης",
                f"Αποτυχία ενημέρωσης γραμμής παραγωγού.\n\n{str(e)}"
            )

    def filter_afm(self, text):
        
        """Φίλτρο για τη στήλη ΑΦΜ - αναζήτηση από την αρχή"""
        if text.strip():
            #  Αναζήτηση από την αρχή του ΑΦΜ
            self.proxy.setFilterRegExp(QRegExp(f"^{text.strip()}"))
        else:
            # Κενό → εμφάνιση όλων
            self.proxy.setFilterRegExp(QRegExp(""))

    def clear_filter(self):
        """Καθαρίζει το φίλτρο και εμφανίζει όλες τις γραμμές"""
        self.ui.lineEdit_search.blockSignals(True)
        self.ui.lineEdit_search.clear()
        self.ui.lineEdit_search.blockSignals(False)
        self.proxy.setFilterRegExp(QRegExp(""))


    def delete_btn(self, p_index):
        try:

            if not p_index.isValid():
                return
            
            # 1. Εντοπισμός του ΑΦΜ από το μοντέλο
            source_index = self.proxy.mapToSource(p_index)
            row = source_index.row()
            afm_item = self.model.item(row, 0)
            if not afm_item:
                return   
            deleted_afm = afm_item.text().strip()

                #  Δημιουργία MessageBox
            msgbox = QMessageBox(self)
            msgbox.setWindowTitle("Επιβεβαίωση διαγραφής")
            msgbox.setText(
                f"Με την ενέργεια αυτή θα διαγραφούν όλες οι καταχωρήσεις για τον ΑΦΜ {deleted_afm}.\n\n"
                "Επιθυμείτε να συνεχίσετε;"
            )
            msgbox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            
            #  Αλλαγή κειμένων κουμπιών
            msgbox.button(QMessageBox.Yes).setText("Ναι")
            msgbox.button(QMessageBox.No).setText("Όχι")
            
            #  Εμφάνιση και λήψη απάντησης
            reply = msgbox.exec_()

            if reply == QMessageBox.No:
                return
            
            if delete_producer(deleted_afm):
                # Αν η βάση επέστρεψε True, αφαιρούμε τη γραμμή από το UI
                self.model.removeRow(row)
                show_temp_message(self, f"Ο ΑΦΜ {deleted_afm} διαγράφηκε οριστικά.", bg_color="#4ABD56")
                self.clear_filter()
                
                
                # Αν ο διαγραμμένος ΑΦΜ είναι αυτός που φαίνεται στο MainWindow, καθαρίζουμε το UI
                if self.main_window:
                    current_afm = self.main_window.ui.lineEdit_afm.text().strip()
                    if deleted_afm == current_afm:
                        self.main_window.clear_ui()
                        self.main_window.has_unsaved_changes = False
                        
            else:
                show_temp_message(self, "Σφάλμα: Η διαγραφή από τη βάση δεδομένων απέτυχε!")
        except Exception as e:
            QMessageBox.critical(
                self,
                "Σφάλμα",
                f"Σφάλμα κατά τη διαγραφή:\n{str(e)}"
            )

    def edit_btn(self, p_index):
        """Λειτουργία για το Μολύβι: Φόρτωση δεδομένων για επεξεργασία"""
        if not p_index.isValid():
            return
        
        # 1. Μετατροπή proxy index σε source index
        source_index = self.proxy.mapToSource(p_index)
        row = source_index.row()
        
        # 2. Λήψη στοιχείων από τη γραμμή του πίνακα
        afm_item = self.model.item(row, 0)
        
        if not afm_item:
            return
        
        afm = afm_item.text().strip()

        #  Έλεγχος μόνο αν πηγαίνουμε σε ΔΙΑΦΟΡΕΤΙΚΟ ΑΦΜ
        if not self.main_window:
            return
        
        current_afm = self.main_window.ui.lineEdit_afm.text().strip()

        if current_afm == afm:
            self.main_window.ui.stackedWidget.setCurrentWidget(self.main_window.ta_page)
            self.main_window.set_lineEdits_editable(False)
            return
        
        if self.main_window.has_unsaved_changes:
            msgbox = QMessageBox(self)
            msgbox.setWindowTitle("Μη αποθηκευμένες αλλαγές")
            msgbox.setText(
                f"Υπάρχουν μη αποθηκευμένες αλλαγές για τον ΑΦΜ: {current_afm}\n\n"
                "Αν συνεχίσετε, οι αλλαγές θα χαθούν."
            )
            msgbox.setInformativeText("Επιθυμείτε να συνεχίσετε;")
            msgbox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msgbox.button(QMessageBox.Yes).setText("Ναι")
            msgbox.button(QMessageBox.No).setText("Όχι")
            msgbox.setDefaultButton(QMessageBox.No)
            
            reply = msgbox.exec_()
        
            if reply == QMessageBox.No:
                return  # Ματαίωση - παραμένει στην αρχική
        
        # 3. Ενημέρωση του MainWindow, μετάβαση σελίδας και clear
       
        self.main_window.ui.lineEdit_afm.setText(afm)

        self.main_window.handle_search()

        
        self.main_window.search_or_edit_performed = True
        self.main_window._lock_actions_only(True)
        self.main_window.ui.savebtn.setEnabled(True)
        self.main_window.ui.exportbtn.setEnabled(True)
        self.main_window.ui.importbtn.setEnabled(False)

        self.main_window.ui.stackedWidget.setCurrentWidget(self.main_window.ta_page)
        self.main_window.ui.searchbtn.setEnabled(False)
        self.main_window.set_lineEdits_editable(False)
        self.clear_filter()

    
    def export_table_to_xlsx(self):
        """Εξαγωγή πίνακα εγγραφών σε xlsx"""
        if self.model.rowCount() == 0:
            show_temp_message(self, "Δεν υπάρχουν εγγραφές για εξαγωγή.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Αποθήκευση αρχείου", "Πίνακας Εγγραφών", "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Εγγραφές"

            # Ημερομηνία & ώρα στο A1
            ws.cell(row=1, column=1, value="Ημερομηνία εκτύπωσης: " + datetime.now().strftime("%d/%m/%Y %H:%M"))

            # Headers (γραμμή 3)
            headers = ["ΑΦΜ", "Όνομα", "Επώνυμο", "Περιφέρεια", "Αρχική ΤΑ", "Μελλοντική ΤΑ", "Μόρια", "Επιλεξιμότητα", "Τελευταία Επεξεργασία"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=3, column=col, value=header)

            # Στήλες που αποθηκεύονται ως αριθμός (0-based index)
            numeric_cols = {4, 5, 6}  # Αρχική ΤΑ, Μελλοντική ΤΑ, Μόρια

            # Δεδομένα (από γραμμή 4)
            for row in range(self.model.rowCount()):
                for col in range(9):
                    item = self.model.item(row, col)
                    text = item.text() if item else ""
                    if col in numeric_cols and text:
                        try:
                            cell = ws.cell(row=row + 4, column=col + 1, value=round(float(text), 2))
                            cell.number_format = "0.00"
                        except ValueError:
                            ws.cell(row=row + 4, column=col + 1, value=text)
                    else:
                        ws.cell(row=row + 4, column=col + 1, value=text)

            # Bold headers (γραμμή 3)
            bold = Font(bold=True)
            for col in range(1, len(headers) + 1):
                ws.cell(row=3, column=col).font = bold

            # Autofit στήλες
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        cell_len = len(str(cell.value)) if cell.value is not None else 0
                        if cell_len > max_length:
                            max_length = cell_len
                    except Exception:
                        pass # nosec B110 - cell width calc, cosmetic only
                ws.column_dimensions[col_letter].width = max_length + 2

            wb.save(file_path)
            show_temp_message(self, "Η εξαγωγή ολοκληρώθηκε!", bg_color="#4ABD56")
        except PermissionError:
            QMessageBox.critical(
                self,
                "Σφάλμα Δικαιωμάτων",
                "Δεν υπάρχουν δικαιώματα εγγραφής στον επιλεγμένο φάκελο.\n\n"
                "Επιλέξτε διαφορετική τοποθεσία."
            )
        except Exception as e:
            QMessageBox.critical(self, "Σφάλμα", f"Σφάλμα κατά την εξαγωγή:\n{str(e)}")

    def _setup_table_windows_compatibility_arxiki(self):
        """Ρυθμίσεις για Windows 7+ compatibility"""

        #  Ενεργοποίηση gridlines (για να φαίνονται)
        self.ui.tableView.setShowGrid(True)

        # Mouse tracking για hover state στα painted delegate buttons (cols 9, 10)
        self.ui.tableView.setMouseTracking(True)

        
        # Vertical header setup
        v_header = self.ui.tableView.verticalHeader()
        v_header.setVisible(True)  # Αν θες να φαίνονται οι αριθμοί
        v_header.setStretchLastSection(False)
        v_header.setDefaultSectionSize(35)
        
        
        self.ui.tableView.setColumnWidth(1, 120)
        self.ui.tableView.setColumnWidth(2, 120)
        self.ui.tableView.setColumnWidth(5, 130)
        self.ui.tableView.setColumnWidth(7, 150)
        self.ui.tableView.setColumnWidth(8, 160)
        

        # Smooth scrolling
        self.ui.tableView.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.ui.tableView.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)

    

    

    
        