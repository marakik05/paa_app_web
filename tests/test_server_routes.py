"""
Integration tests για τα API routes του server.py (Flask test_client).

Η βάση ανακατευθύνεται σε προσωρινό αρχείο (μέσω database_manager.DB_PATH) ΠΡΙΝ
γίνει το πρώτο import του server.py, ώστε το setup_database() που τρέχει στο
module-level του server.py (server.py:8) να γράψει στο temp αρχείο και όχι στην
πραγματική βάση χρήστη στο %LOCALAPPDATA%.
"""
import os
import io
import sys
import sqlite3
import tempfile
import shutil
import unittest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import database_manager as db

_tmp_dir = None
_orig_db_path = None
app = None
server = None


def setUpModule():
    global _tmp_dir, _orig_db_path, app, server
    _tmp_dir = tempfile.mkdtemp()
    _orig_db_path = db.DB_PATH
    db.DB_PATH = os.path.join(_tmp_dir, 'test.db')
    db.setup_database()  # ρητή δημιουργία πινάκων — δεν αρκεί το module-level setup_database()
                          # του server.py αν το 'server' module είναι ήδη cached από άλλο test file

    import server as server_module
    server = server_module
    app = server.app
    app.testing = True


def tearDownModule():
    db.DB_PATH = _orig_db_path
    shutil.rmtree(_tmp_dir, ignore_errors=True)


def _clear_db():
    conn = sqlite3.connect(db.DB_PATH)
    cur = conn.cursor()
    cur.execute("DELETE FROM osde_entries")
    cur.execute("DELETE FROM producers")
    conn.commit()
    conn.close()


def _write_xlsx_bytes(rows, sheet_name="Τυπική Απόδοση", headers=None):
    import openpyxl
    default_headers = [
        "ΑΦΜ", "Όνομα", "Επώνυμο", "Περιφέρεια",
        "Κατηγορία ΟΣΔΕ", "Περιγραφή",
        "Έκταση/Αριθμός ζώων", "Βιολογικά",
        "Δένδρα >=4 ετών", "Δένδρα <4 ετών", "Αμπέλι >3 ετών",
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers if headers is not None else default_headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


class _ClientMixin(unittest.TestCase):

    def setUp(self):
        _clear_db()
        self.client = app.test_client()


# ─── /api/regions ─────────────────────────────────────────────────────

class TestRegions(_ClientMixin):

    def test_returns_periferies_list(self):
        resp = self.client.get('/api/regions')
        self.assertEqual(resp.status_code, 200)
        data = resp.get_json()
        self.assertIsInstance(data, list)
        self.assertEqual(data[0], "--Επιλέξτε")
        self.assertIn("Αττική", data)


# ─── /api/producer/<afm>/exists ──────────────────────────────────────

class TestProducerExists(_ClientMixin):

    def test_nonexistent_afm(self):
        resp = self.client.get('/api/producer/000000000/exists')
        self.assertEqual(resp.get_json(), {'exists': False})

    def test_existing_afm(self):
        db.save_producer_basics('123456789', 'A', 'B', 'Αττική')
        resp = self.client.get('/api/producer/123456789/exists')
        self.assertEqual(resp.get_json(), {'exists': True})


# ─── /api/producer/<afm>/full ────────────────────────────────────────

class TestProducerFull(_ClientMixin):

    def test_not_found(self):
        resp = self.client.get('/api/producer/000000000/full')
        self.assertEqual(resp.get_json(), {'found': False})

    def test_found_with_rows(self):
        db.save_producer_basics('123456789', 'Γιάννης', 'Παπαδόπουλος', 'Αττική')
        db.save_scenario_data('123456789', 'initial', [
            ('123456789', 'initial', 'CAT', 'DESC', 100.0, 5.0,
             'Συμβατικά', 0, 0, '', 500.0, 500.0, 500.0, 0, 0, 0)
        ])
        resp = self.client.get('/api/producer/123456789/full')
        data = resp.get_json()
        self.assertTrue(data['found'])
        self.assertEqual(data['name'], 'Γιάννης')
        self.assertEqual(data['surname'], 'Παπαδόπουλος')
        self.assertEqual(data['region'], 'Αττική')
        self.assertEqual(len(data['initial_rows']), 1)


# ─── POST /api/producer/<afm>/save ───────────────────────────────────

class TestSaveProducer(_ClientMixin):

    def test_saves_basics_and_rows(self):
        body = {
            'name': 'Μαρία', 'surname': 'Κ', 'region': 'Κρήτη',
            'initial_rows': [
                ['CAT', 'DESC', 100.0, 5.0, 'Συμβατικά', 0, 0, '', 500.0, 500.0, 500.0, 0, 0, 0]
            ],
        }
        resp = self.client.post('/api/producer/123456789/save', json=body)
        self.assertEqual(resp.get_json(), {'ok': True})

        producer = db.fetch_producer('123456789')
        self.assertEqual(producer, ('Μαρία', 'Κ', 'Κρήτη'))
        entries = db.fetch_entries('123456789', 'initial')
        self.assertEqual(len(entries), 1)

    def test_save_without_rows_does_not_touch_entries(self):
        db.save_producer_basics('123456789', 'Old', 'Name', 'Αττική')
        db.save_scenario_data('123456789', 'initial', [
            ('123456789', 'initial', 'CAT', 'DESC', 100.0, 5.0, '', 0, 0, '', 500.0, 500.0, 500.0, 0, 0, 0)
        ])
        body = {'name': 'New', 'surname': 'Name', 'region': 'Αττική'}
        resp = self.client.post('/api/producer/123456789/save', json=body)
        self.assertEqual(resp.get_json(), {'ok': True})
        # initial_rows not in body → entries unchanged
        self.assertEqual(len(db.fetch_entries('123456789', 'initial')), 1)


# ─── /api/producers ───────────────────────────────────────────────────

class TestProducersList(_ClientMixin):

    def test_empty_list(self):
        resp = self.client.get('/api/producers')
        self.assertEqual(resp.get_json(), [])

    def test_returns_all_producers(self):
        db.save_producer_basics('111111111', 'A', 'A', 'Αττική')
        db.save_producer_basics('222222222', 'B', 'B', 'Κρήτη')
        resp = self.client.get('/api/producers')
        data = resp.get_json()
        afms = [r[0] for r in data]
        self.assertEqual(afms, ['111111111', '222222222'])  # ORDER BY afm ASC


# ─── DELETE /api/producer/<afm> ──────────────────────────────────────

class TestDeleteProducer(_ClientMixin):

    def test_deletes_existing_producer(self):
        db.save_producer_basics('123456789', 'A', 'B', 'Αττική')
        resp = self.client.delete('/api/producer/123456789')
        self.assertEqual(resp.get_json(), {'ok': True})
        self.assertIsNone(db.fetch_producer('123456789'))

    def test_deleting_nonexistent_still_ok(self):
        resp = self.client.delete('/api/producer/999999999')
        self.assertEqual(resp.get_json(), {'ok': True})


# ─── /api/ta/reference ────────────────────────────────────────────────

class TestTaReference(_ClientMixin):

    def test_returns_mapping(self):
        resp = self.client.get('/api/ta/reference')
        data = resp.get_json()
        self.assertIn('mapping', data)
        self.assertIsInstance(data['mapping'], dict)


# ─── POST /api/ta/recalculate ─────────────────────────────────────────

class TestTaRecalculate(_ClientMixin):

    def test_recalculate_structure(self):
        body = {
            'region': 'Αττική',
            'rows': [
                {'category': 'ΣΚΛΗΡΟΣ ΣΙΤΟΣ', 'description': 'Άγνωστη Ποικιλία',
                 'quantity': '5', 'trees_over_4': '', 'trees_under_4': '', 'vine_over_3': ''}
            ],
        }
        resp = self.client.post('/api/ta/recalculate', json=body)
        data = resp.get_json()
        self.assertEqual(len(data['rows']), 1)
        row = data['rows'][0]
        for key in ('typical_output', 'output_per_choice', 'lock_ampeli', 'lock_trees'):
            self.assertIn(key, row)
        for key in ('total_output', 'ta_productive', 'ta_plant', 'ta_animal', 'ta_bees'):
            self.assertIn(key, data['totals'])

    def test_recalculate_empty_rows(self):
        resp = self.client.post('/api/ta/recalculate', json={'region': 'Αττική', 'rows': []})
        data = resp.get_json()
        self.assertEqual(data['rows'], [])
        self.assertIsNone(data['totals']['total_output'])


# ─── POST /api/import/parse ───────────────────────────────────────────

class TestImportParse(_ClientMixin):

    def test_no_file_returns_error(self):
        resp = self.client.post('/api/import/parse', data={})
        data = resp.get_json()
        self.assertFalse(data['ok'])
        self.assertIn('error', data)

    def test_invalid_file_returns_error_not_500(self):
        """Magic-bytes check αποτυγχάνει → ok:False, ΟΧΙ 500 (ValueError caught)."""
        bad = io.BytesIO(b"NOT AN XLSX FILE")
        resp = self.client.post(
            '/api/import/parse',
            data={'file': (bad, 'bad.xlsx')},
            content_type='multipart/form-data',
        )
        self.assertEqual(resp.status_code, 200)
        data = resp.get_json()
        self.assertFalse(data['ok'])

    def test_valid_file_new_producer(self):
        xlsx_bytes = _write_xlsx_bytes([
            ["123456789", "Maria", "K", "Αττική", "Cat", "Desc", 1, "Συμβατικά", "", "", ""]
        ])
        resp = self.client.post(
            '/api/import/parse',
            data={'file': (io.BytesIO(xlsx_bytes), 'data.xlsx')},
            content_type='multipart/form-data',
        )
        data = resp.get_json()
        self.assertTrue(data['ok'])
        self.assertEqual(len(data['new_data']), 1)
        self.assertEqual(data['new_data'][0]['afm'], '123456789')
        self.assertEqual(data['conflicts'], [])

    def test_valid_file_existing_afm_is_conflict(self):
        db.save_producer_basics('123456789', 'Old', 'Name', 'Αττική')
        xlsx_bytes = _write_xlsx_bytes([
            ["123456789", "Maria", "K", "Αττική", "Cat", "Desc", 1, "Συμβατικά", "", "", ""]
        ])
        resp = self.client.post(
            '/api/import/parse',
            data={'file': (io.BytesIO(xlsx_bytes), 'data.xlsx')},
            content_type='multipart/form-data',
        )
        data = resp.get_json()
        self.assertTrue(data['ok'])
        self.assertEqual(data['new_data'], [])
        self.assertEqual(len(data['conflicts']), 1)
        self.assertEqual(data['conflicts'][0]['afm'], '123456789')

    def test_invalid_region_in_file_returns_error(self):
        xlsx_bytes = _write_xlsx_bytes([
            ["123456789", "Maria", "K", "Ανύπαρκτη", "Cat", "Desc", 1, "Συμβατικά", "", "", ""]
        ])
        resp = self.client.post(
            '/api/import/parse',
            data={'file': (io.BytesIO(xlsx_bytes), 'data.xlsx')},
            content_type='multipart/form-data',
        )
        data = resp.get_json()
        self.assertFalse(data['ok'])
        self.assertIn('error', data)


# ─── POST /api/import/execute ─────────────────────────────────────────

class TestImportExecute(_ClientMixin):

    def test_execute_new_producer(self):
        body = {
            'producers': [{
                'afm': '123456789', 'name': 'Maria', 'surname': 'K', 'region': 'Αττική',
                'rows': [{'category_osde': 'Cat', 'description': 'Desc', 'quantity': '1',
                          'certification': '', 'trees_over_4': '', 'trees_under_4': '', 'vine_over_3': ''}],
            }],
            'decisions': {},
        }
        resp = self.client.post('/api/import/execute', json=body)
        data = resp.get_json()
        self.assertTrue(data['ok'])
        self.assertEqual(data['total_success'], 1)
        self.assertIsNotNone(db.fetch_producer('123456789'))

    def test_execute_conflict_skip(self):
        db.save_producer_basics('123456789', 'Old', 'Name', 'Αττική')
        body = {
            'producers': [{
                'afm': '123456789', '_conflict': True, 'name': 'New', 'surname': 'Name',
                'region': 'Κρήτη', 'rows': [],
            }],
            'decisions': {'123456789': 'skip'},
        }
        resp = self.client.post('/api/import/execute', json=body)
        data = resp.get_json()
        self.assertTrue(data['ok'])
        self.assertEqual(data['total_success'], 0)
        # Skipped → producer basics unchanged
        producer = db.fetch_producer('123456789')
        self.assertEqual(producer[0], 'Old')

    def test_execute_conflict_replace(self):
        db.save_producer_basics('123456789', 'Old', 'Name', 'Αττική')
        body = {
            'producers': [{
                'afm': '123456789', '_conflict': True, 'name': 'New', 'surname': 'Name',
                'region': 'Κρήτη', 'rows': [],
            }],
            'decisions': {'123456789': 'replace'},
        }
        resp = self.client.post('/api/import/execute', json=body)
        data = resp.get_json()
        self.assertTrue(data['ok'])
        self.assertEqual(data['total_success'], 1)
        producer = db.fetch_producer('123456789')
        self.assertEqual(producer[0], 'New')
        self.assertEqual(producer[2], 'Κρήτη')


# ─── POST /api/producer/<afm>/export ──────────────────────────────────

class TestExportProducer(_ClientMixin):

    def test_export_returns_xlsx_reading_from_body_not_db(self):
        """Export διαβάζει ΠΑΝΤΑ από το request body — όχι από τη βάση."""
        import openpyxl

        # DB έχει διαφορετικά δεδομένα από αυτά που θα σταλούν στο export
        db.save_producer_basics('123456789', 'InDB', 'InDB', 'Αττική')

        body = {
            'name': 'FromBody', 'surname': 'FromBody', 'region': 'Κρήτη',
            'rows': [
                ['CAT', 'DESC', 100.0, 5.0, 'Συμβατικά', 0, 0, '', 500.0,
                 500.0, 500.0, 500.0, 0, 0]
            ],
            'totals': {'total_ta': 500.0, 'ta_prod': 500.0, 'ta_plant': 500.0,
                       'ta_animal': 0, 'ta_bees': 0},
        }
        resp = self.client.post('/api/producer/123456789/export', json=body)
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheetml', resp.content_type)

        wb = openpyxl.load_workbook(io.BytesIO(resp.data))
        ws = wb["Τυπική Απόδοση"]
        data_row = [c.value for c in ws[2]]
        self.assertEqual(data_row[0], '123456789')   # afm (from URL)
        self.assertEqual(data_row[1], 'FromBody')     # name (from body, not DB)
        self.assertEqual(data_row[3], 'Κρήτη')         # region (from body, not DB)


if __name__ == '__main__':
    unittest.main()
