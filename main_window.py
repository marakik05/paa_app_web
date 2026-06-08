from PySide2.QtWidgets import QMainWindow, QFileDialog, QMessageBox, QDialog, QProgressDialog, QApplication
from PySide2.QtGui import QRegExpValidator, QPixmap, QTextDocument
from PySide2.QtCore import QRegExp, Qt, QEvent, QTimer
from ui.ui_window import Ui_MainWindow
from pages.page_0_ta import taPage
from pages.page_1_epileximotita import epileximotitaPage
from pages.page_mellontiki import mellontikiPage
from pages.page_arxiki import arxikiPage
from pages.page_moria import moriaPage
from utils.excel_loader import PERIFERIES, LOCK_AMPELI_NORM, in_norm_set, norm, NORM_YES, resource_path, DEFAULT
import openpyxl
from openpyxl.styles import Font
from database_manager import fetch_moria, save_moria_data, import_producers_batch_transaction, save_producer_basics, save_scenario_data, save_eligibility_data, fetch_producer, fetch_entries, fetch_eligibility, setup_database, fetch_all_producers
from utils.message import show_temp_message_main, ImportConflictDialog, MissingCategoriesDialog
import csv


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        setup_database()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        #Εισαγωγή logo
        pixmap = QPixmap(resource_path("logo/gaia.png"))
        pixmap = pixmap.scaled(80, 40, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        self.ui.label_logo.setPixmap(pixmap)
        self.ui.label_logo.setAlignment(Qt.AlignRight)
        #Τίτλος
        self.setWindowTitle("Παρέμβαση Π3-73-2.1-Σχέδια Βελτίωσης")
        
        #Τιμές combobox για την Περιφέρεια 
        self.ui.combo_periferia.addItems(PERIFERIES)      
        
        #Σύνδεση των ui pages με τις αντίστοιχες κλάσεις της κάθε σελίδας
        self.ta_page = taPage()
        self.epi_page = epileximotitaPage()
        self.mel_page=mellontikiPage()
        self.arx_page=arxikiPage(main_window=self)
        self.moria_page=moriaPage()

        # Cache categories από το ήδη φορτωμένο value_mapping
        self._ta_categories_cache = {
            (norm(k[0]), norm(k[1]))
            for k in self.ta_page.value_mapping.keys()
}

        # Caches για fast lookup κατά το import validation
        self._valid_cats_cache = {norm(k) for k in self.ta_page.mapping.keys()}
        self._valid_descs_cache = {
            norm(v) for varieties in self.ta_page.mapping.values() for v in varieties
        }

        # Reverse lookup: norm(cat) → canonical cat (από ta.xlsx)
        self._canonical_cat_by_norm = {
            norm(k): k for k in self.ta_page.mapping.keys()
        }

        # Reverse lookup: (norm(cat), norm(desc)) → (canonical_cat, canonical_desc)
        self._canonical_pair_by_norm = {
            (norm(cat), norm(desc)): (cat, desc)
            for cat, desc in self.ta_page.value_mapping.keys()
        }

        #Σύνδεση της lineEdit_7 με τη συνολική ΤΑ του πίνακα ΤΑ-αρχικής
        self.ta_page.totalChanged.connect(self.epi_page.update_lineEdit_7)
        self.ta_page.totalChanged.connect(self.moria_page.update_lineEdit_2_1) 
        self.ta_page.totalChanged.connect(self.moria_page.update_lineEdit_2_2) 
        self.ta_page.biologicChanged.connect(self.moria_page.update_lineEdit_1_2)            

        #Προσθήκη αρχικής σελίδας του designer
        self.ui.stackedWidget.setCurrentIndex(0)
         # Προσθήκη στο stackedWidget των κλάσεων των σελίδων
        self.ui.stackedWidget.addWidget(self.ta_page)
        self.ui.stackedWidget.addWidget(self.epi_page)
        self.ui.stackedWidget.addWidget(self.mel_page)
        self.ui.stackedWidget.addWidget(self.arx_page)
        self.ui.stackedWidget.addWidget(self.moria_page)

        #  default page όταν ανοίγει η εφαρμογή
        self.ui.stackedWidget.setCurrentWidget(self.arx_page)

        # Mήνυμα στην ta_page αν δεν έχουν βρεθεί καλλιέργειες/ποικιλίες στη βάση
        self.ui.stackedWidget.currentChanged.connect(self._on_current_page_changed)
                 
        # Σύνδεση των pages με QAction και απενεργοποίηση searchbtn
        self.ui.ta.triggered.connect(
            lambda: (self.ui.stackedWidget.setCurrentWidget(self.ta_page),
            self.ui.searchbtn.setEnabled(False))
        )
        
        self.ui.epileximotita.triggered.connect(
           lambda: (self.ui.stackedWidget.setCurrentWidget(self.epi_page),
           self.ui.searchbtn.setEnabled(False))
        )

        self.ui.mellontiki.triggered.connect(
           lambda: (self.ui.stackedWidget.setCurrentWidget(self.mel_page),
           self.ui.searchbtn.setEnabled(False))
        )

        self.ui.arxiki.triggered.connect(
           lambda: (self.ui.stackedWidget.setCurrentWidget(self.arx_page),
           self.ui.searchbtn.setEnabled(True))
        )

        self.ui.moria.triggered.connect(
          lambda: self.ui.stackedWidget.setCurrentWidget(self.moria_page)
        )

        self.ui.combo_periferia.currentTextChanged.connect(self.on_periferia_changed)
        
        #περιορισμός έως 9 ψηφίων στο lineEdit του ΑΦΜ
        
        regex = QRegExp(r"\d{0,9}")  # από 0 έως 9 ψηφία
        validator = QRegExpValidator(regex)
        self.ui.lineEdit_afm.setValidator(validator)

        # Σύνδεση κουμπιών με τις συναρτήσεις της βάσης
        self.ui.savebtn.clicked.connect(self.handle_save)
        self.ui.searchbtn.clicked.connect(self.handle_search)

        # Παρακολούθηση αλλαγών στο ΑΦΜ για κλείδωμα QActions
        self.ui.lineEdit_afm.textChanged.connect(self.on_afm_changed)

        # Αν ο χρήστης σβήσει το ΑΦΜ, καθάρισε όλη την οθόνη
        self.ui.lineEdit_afm.textChanged.connect(
            lambda text: self.clear_ui() if not text.strip() else None
        )

        # Σύνδεση του 2ου πίνακα με τον 1ο
        self.mel_page.set_source_table(self.ta_page)

        #σύνδεση με το κουμπί εξαγωγή σε excel
        self.ui.exportbtn.clicked.connect(self.export_table_to_excel)
        #σύνδεση με κουμπί import
        self.ui.importbtn.clicked.connect(self.import_data_from_file)
        #σύνδεση qaction με τις def go_to

        self.ui.arxiki.triggered.connect(self.go_to_arxiki)
        self.ui.ta.triggered.connect(self.go_to_ta)
        self.ui.mellontiki.triggered.connect(self.go_to_ta_mel)
        self.ui.epileximotita.triggered.connect(self.go_to_epi)
        self.ui.moria.triggered.connect(self.go_to_moria)

                 

        # Παρακολούθηση αλλαγών στους πίνακες Αρχικής-Μελλοντικής ΤΑ
        self.ta_page.model.dataChanged.connect(self.mark_as_changed)
        self.mel_page.model.dataChanged.connect(self.mark_as_changed)
        self.ta_page.model.rowsInserted.connect(self.mark_as_changed)
        self.mel_page.model.rowsInserted.connect(self.mark_as_changed)
        self.ta_page.model.rowsRemoved.connect(self.mark_as_changed)
        self.mel_page.model.rowsRemoved.connect(self.mark_as_changed)

        #Παρακολούθηση αλλαγών στα combobox και lineEdits της επιλεξιμότητας
        self.epi_page.ui.lineEdit_9.textChanged.connect(self.mark_as_changed)
        self.epi_page.ui.comboBox_2.currentIndexChanged.connect(self.mark_as_changed)
        self.epi_page.ui.comboBox_3.currentIndexChanged.connect(self.mark_as_changed)
        self.epi_page.ui.comboBox_4.currentIndexChanged.connect(self.mark_as_changed)
        self.epi_page.ui.comboBox_5.currentIndexChanged.connect(self.mark_as_changed)
        self.epi_page.ui.comboBox_6.currentIndexChanged.connect(self.mark_as_changed)  
        self.epi_page.ui.comboBox.currentIndexChanged.connect(self.mark_as_changed)
        self.epi_page.ui.comboBox_7.currentIndexChanged.connect(self.mark_as_changed)
        self.epi_page.ui.comboBox_8.currentIndexChanged.connect(self.mark_as_changed)

        #Παρακολούθηση αλλαγών στα combobox και lineEdits της μοριοδότησης   
        self.moria_page.ui.comboBox_1_1.currentIndexChanged.connect(self.mark_as_changed)  
        self.moria_page.ui.lineEdit_3_1_1.textChanged.connect(self.mark_as_changed)
        self.moria_page.ui.comboBox_3_1_2.currentIndexChanged.connect(self.mark_as_changed) 
        self.moria_page.ui.comboBox_3_1_3.currentIndexChanged.connect(self.mark_as_changed)
        self.moria_page.ui.comboBox_3_1_4.currentIndexChanged.connect(self.mark_as_changed) 
        self.moria_page.ui.comboBox_3_2.currentIndexChanged.connect(self.mark_as_changed)
        self.moria_page.ui.comboBox_3_3.currentIndexChanged.connect(self.mark_as_changed) 
        self.moria_page.ui.comboBox_3_4.currentIndexChanged.connect(self.mark_as_changed)
        self.moria_page.ui.comboBox_3_5.currentIndexChanged.connect(self.mark_as_changed)
        self.moria_page.ui.lineEdit_4_1.textChanged.connect(self.mark_as_changed) 
        self.moria_page.ui.lineEdit_5_1.textChanged.connect(self.mark_as_changed)
        self.moria_page.ui.comboBox_6_1.currentIndexChanged.connect(self.mark_as_changed)
        self.moria_page.ui.comboBox_7_1.currentIndexChanged.connect(self.mark_as_changed)
        self.moria_page.ui.lineEdit_budget.textChanged.connect(self.mark_as_changed)     
            
        # Παρακολούθηση αλλαγών στα LineEdits του MainWindow
       
        self.ui.lineEdit_name.textChanged.connect(self.mark_as_changed)
        self.ui.lineEdit_surname.textChanged.connect(self.mark_as_changed)
        self.ui.combo_periferia.currentIndexChanged.connect(self.mark_as_changed)

        #eventfilter στο lineEdit_afm
        self.ui.lineEdit_afm.installEventFilter(self)
        #flags
        self.has_unsaved_changes = False #  Flag για μη αποθηκευμένες αλλαγές
        self._marking_changes = False  # Αποφυγή πολλαπλών calls από μη αποθηκευμενα αφμ
        self._afm_focus_dialog_open = False
        self._is_closing = False # Flag για κλείσιμο εφαρμογής

        self.search_or_edit_performed = False # Flag για το αν έχει γίνει search ή edit

        self._pending_ta_warning_message = None #για την εμφάνιση μηνύματος στην ta_page

        self._pending_mel_warning_message = None #για την εμφάνιση μηνύματος στην mel_page
       
        self._lock_actions(False) # Αρχικό κλείδωμα qaction (εκτός Αρχικής)

        self._setup_initial_focus()  

        self._setup_minimum_sizes()

       
    def _setup_initial_focus(self):
        """Ρύθμιση focus για καθαρή εμφάνιση κατά το άνοιγμα"""
        # Widgets που είναι disabled στην αρχή - NoFocus
        self.ui.lineEdit_afm.setFocusPolicy(Qt.ClickFocus)
        self.ui.lineEdit_name.setFocusPolicy(Qt.ClickFocus)
        self.ui.lineEdit_surname.setFocusPolicy(Qt.ClickFocus)
        self.ui.combo_periferia.setFocusPolicy(Qt.ClickFocus)
        
       
        # Clear focus από όλα
        self.setFocus()     
        

    def on_afm_changed(self, text):
        """Κλείδωμα/ξεκλείδωμα QActions βάσει μήκους ΑΦΜ"""
        afm = text.strip()
        self.search_or_edit_performed = False
        self.ui.savebtn.setEnabled(False)
        self.ui.exportbtn.setEnabled(False)
        self.ui.importbtn.setEnabled(False)
        if hasattr(self, '_current_loaded_afm') and self._current_loaded_afm:
            if afm != self._current_loaded_afm:
                # ΑΦΜ άλλαξε → καθαρισμός
                
                
                self.clear_ui_not_afm()
                
                self._current_loaded_afm = ""  # Reset
                self.has_unsaved_changes = False
                
                
            
        # Αν κενό → καθάρισε UI
        if not afm:
            self.clear_ui()
            self._lock_actions_only(False)  # Κλείδωμα
            self.ui.importbtn.setEnabled(True)
            return
        
        # Αν 9 ψηφία → ξεκλείδωμα
        if len(afm) < 9:
            self._lock_actions(False)
            self.clear_ui_not_afm()
            return
            
        try:
            from database_manager import fetch_producer
            producer = fetch_producer(afm)
            
            if not producer:
                # ΔΕΝ υπάρχει → Enable για νέα εγγραφή
                self._lock_actions_only(True)
                self.search_or_edit_performed = True
                self.ui.savebtn.setEnabled(True)
                self.ui.exportbtn.setEnabled(True)
                

                self.clear_ui_not_afm()
                
            else:
                self._lock_actions_only(False)
                self.ui.savebtn.setEnabled(False)
                self.ui.exportbtn.setEnabled(False)
                
                self.clear_ui_not_afm()
                
        except Exception:
            # Σφάλμα βάσης → Ignore
            self._lock_actions_only(True)
            self.search_or_edit_performed = True
            self.ui.savebtn.setEnabled(True)
            self.ui.exportbtn.setEnabled(True)
            self.ui.importbtn.setEnabled(True)
            self.clear_ui_not_afm()
        

  

    def _lock_actions_only(self, enabled):
        """Ενεργοποίηση/Απενεργοποίηση ΜΟΝΟ των QActions (ΟΧΙ των κουμπιών)"""
        # Η "Αρχική" είναι ΠΑΝΤΑ ενεργή
        self.ui.arxiki.setEnabled(True)
        
        # Οι υπόλοιπες ενεργοποιούνται μόνο με 9 ψηφία
        self.ui.ta.setEnabled(enabled)
        self.ui.mellontiki.setEnabled(enabled)
        self.ui.epileximotita.setEnabled(enabled)
        self.ui.moria.setEnabled(enabled)

    def _lock_actions(self, enabled):
            """Ενεργοποίηση/Απενεργοποίηση QActions ΚΑΙ κουμπιών"""
            # QActions
            self._lock_actions_only(enabled)
            
            # Κουμπιά Save/Export
            self.ui.savebtn.setEnabled(enabled)
            self.ui.exportbtn.setEnabled(enabled)
        
    
    def eventFilter(self, obj, event):
        """Προειδοποιητικό μήνυμα μη αποθηκευμένων αλλαγών στο lineEdit_afm"""
        if obj == self.ui.lineEdit_afm:            
            
            # Όταν κάνει κλικ (focus in)
            if event.type() == QEvent.FocusIn:
                
                
                if self.ui.stackedWidget.currentWidget() != self.arx_page:
                    return super().eventFilter(obj, event)
                if self._is_closing:
                    return super().eventFilter(obj, event)
                if self._afm_focus_dialog_open:
                    return super().eventFilter(obj, event)               
                
                
                if self.has_unsaved_changes:
                    self._afm_focus_dialog_open=True

                    msgbox = QMessageBox(self)
                    msgbox.setWindowTitle("Μη αποθηκευμένες αλλαγές")
                    msgbox.setText("Υπάρχουν μη αποθηκευμένες αλλαγές!\n\n"
                    "Αποθηκεύστε πρώτα τις αλλαγές και μετά αλλάξτε ΑΦΜ.")
                    msgbox.setStandardButtons(QMessageBox.Ok)
                    msgbox.button(QMessageBox.Ok).setText("Εντάξει")             
                    msgbox.exec_()                    
                    self._afm_focus_dialog_open = False 
                   
                    return True
                
            return super().eventFilter(obj, event)
        
    

    
    def mark_as_changed(self):
        """Σηματοδοτεί ότι υπάρχουν μη αποθηκευμένες αλλαγές"""
        # Αποφυγή πολλαπλών εκτελέσεων
        if self._marking_changes:
            return
        
        self._marking_changes = True
        
        # Έλεγχος αν υπάρχει έγκυρος ΑΦΜ
        afm = self.ui.lineEdit_afm.text().strip()
        if len(afm) == 9:
            self.has_unsaved_changes = True
        
        self._marking_changes = False

    def closeEvent(self, event):
        """Μήνυμα μη αποθηκευμένων αλλαγών όταν κλείνει η εφαρμογή"""
        self._is_closing = True
        #  Έλεγχος για μη αποθηκευμένες αλλαγές
        if self.has_unsaved_changes:
            msgbox = QMessageBox(self)
            msgbox.setWindowTitle("Μη αποθηκευμένες αλλαγές")
            msgbox.setText("Υπάρχουν μη αποθηκευμένες αλλαγές!")
            msgbox.setInformativeText("Επιθυμείτε να κλείσετε χωρίς αποθήκευση;")
            msgbox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msgbox.button(QMessageBox.Yes).setText("Ναι")
            msgbox.button(QMessageBox.No).setText("Όχι")
            msgbox.setDefaultButton(QMessageBox.No)

            reply = msgbox.exec_()
            
            if reply == QMessageBox.No:
                self._is_closing = False
                event.ignore()  # Ματαίωση κλεισίματος
                return           
        
        event.accept()

    def _block_all_signals(self, block):
        """Block/Unblock όλα τα signals"""
        #main-window
        self.ui.lineEdit_name.blockSignals(block)
        self.ui.lineEdit_surname.blockSignals(block)
        self.ui.combo_periferia.blockSignals(block)
        #πίνακες
        self.ta_page.model.blockSignals(block)
        self.mel_page.model.blockSignals(block)
        #σελίδα επιλεξιμότητας
        self.epi_page.ui.lineEdit_9.blockSignals(block)
        self.epi_page.ui.comboBox_2.blockSignals(block)
        self.epi_page.ui.comboBox_3.blockSignals(block)
        self.epi_page.ui.comboBox_4.blockSignals(block)
        self.epi_page.ui.comboBox_5.blockSignals(block)
        self.epi_page.ui.comboBox_6.blockSignals(block) 
        self.epi_page.ui.comboBox.blockSignals(block) 
        self.epi_page.ui.comboBox_7.blockSignals(block) 
        self.epi_page.ui.comboBox_8.blockSignals(block) 
        #σελίδα μοριοδότησης  
        self.moria_page.ui.comboBox_1_1.blockSignals(block)  
        self.moria_page.ui.lineEdit_3_1_1.blockSignals(block)
        self.moria_page.ui.comboBox_3_1_2.blockSignals(block) 
        self.moria_page.ui.comboBox_3_1_3.blockSignals(block) 
        self.moria_page.ui.comboBox_3_1_4.blockSignals(block)
        self.moria_page.ui.comboBox_3_2.blockSignals(block)
        self.moria_page.ui.comboBox_3_3.blockSignals(block) 
        self.moria_page.ui.comboBox_3_4.blockSignals(block)
        self.moria_page.ui.comboBox_3_5.blockSignals(block) 
        self.moria_page.ui.lineEdit_4_1.blockSignals(block) 
        self.moria_page.ui.lineEdit_5_1.blockSignals(block) 
        self.moria_page.ui.comboBox_6_1.blockSignals(block)
        self.moria_page.ui.comboBox_7_1.blockSignals(block)
        self.moria_page.ui.lineEdit_budget.blockSignals(block)     

    
    def set_lineEdits_editable(self, editable):
        """Κάνει τα lineEdits editable ή non-editable"""
        self.ui.lineEdit_afm.setReadOnly(not editable)
        self.ui.lineEdit_name.setReadOnly(not editable)
        self.ui.lineEdit_surname.setReadOnly(not editable)
        
        
    
    def go_to_arxiki(self):
        self.ui.stackedWidget.setCurrentWidget(self.arx_page)
        self.set_lineEdits_editable(True) 
        self.ui.searchbtn.setEnabled(True)
        self.arx_page.clear_filter() # Καθαρισμός του search filter
        self.ui.lineEdit_afm.clearFocus()#για να μην γίνεται focus το LineEdit του ΑΦΜ
        self.ui.lineEdit_name.clearFocus()
        self.ui.lineEdit_surname.clearFocus() 
        

        
    
    def go_to_ta(self):
        self.ui.stackedWidget.setCurrentWidget(self.ta_page)
        self._show_pending_ta_warning()
        self.set_lineEdits_editable(False)
        self.ui.searchbtn.setEnabled(False)
        self.ui.importbtn.setEnabled(False)
        self.ui.lineEdit_afm.clearFocus()
        self.ui.lineEdit_name.clearFocus()
        self.ui.lineEdit_surname.clearFocus()


    def go_to_ta_mel(self):
        self.ui.stackedWidget.setCurrentWidget(self.mel_page)
        self._show_pending_mel_warning()
        self.set_lineEdits_editable(False)
        self.ui.searchbtn.setEnabled(False)
        self.ui.importbtn.setEnabled(False)
        self.ui.lineEdit_afm.clearFocus()
        self.ui.lineEdit_name.clearFocus()
        self.ui.lineEdit_surname.clearFocus()


    def go_to_epi(self):
        self.ui.stackedWidget.setCurrentWidget(self.epi_page)
        self.set_lineEdits_editable(False)
        self.ui.searchbtn.setEnabled(False)
        self.ui.importbtn.setEnabled(False)
        self.ui.lineEdit_afm.clearFocus()
        self.ui.lineEdit_name.clearFocus()
        self.ui.lineEdit_surname.clearFocus()

    def go_to_moria(self):
        self.ui.stackedWidget.setCurrentWidget(self.moria_page)
        self.set_lineEdits_editable(False)
        self.ui.searchbtn.setEnabled(False)
        self.ui.importbtn.setEnabled(False)
        self.ui.lineEdit_afm.clearFocus()
        self.ui.lineEdit_name.clearFocus()
        self.ui.lineEdit_surname.clearFocus()
        

          
    def on_periferia_changed(self, text):
        """def για τις αλλαγές στο combobox της περιφέρειας και τον επανυπολογισμό των τιμών"""
       
    # Ενημέρωση πίνακα Αρχικής ΤΑ
        self.ta_page.selected_periferia = text
        if self.ta_page.model.rowCount() > 0:  
            for row in range(self.ta_page.model.rowCount()):
                self.ta_page.recalculate(row)            
        
        # Ενημέρωση πίνακα Μελλοντικής ΤΑ
        self.mel_page.selected_periferia = text
        if self.mel_page.model.rowCount() > 0: 
            for row in range(self.mel_page.model.rowCount()):
                self.mel_page.recalculate(row) 
        # Ενημέρωση σελίδας επιλεξιμότητας
        self.epi_page.selected_periferia=text
        self.epi_page.update_lineEdit_8()

        # Ενημέρωση σελίδας μοριοδότησης
        self.moria_page.selected_periferia=text
        
    
    def export_table_to_excel(self):   
        """def για εξαγωγή σε excel"""
        def label_text(label):
            #Βοηθητική συνάρτηση για αφαιρεση του html text από τα label οταν εξάγονται σε excel
            doc = QTextDocument()
            doc.setHtml(label.text())
            return doc.toPlainText()
    
        def to_float_or_none(text):
            try:
                return round(float(text.replace(",", ".")), 2)
            except (ValueError, TypeError, AttributeError):
                return None
        try:
            afm = self.ui.lineEdit_afm.text()
            surname = self.ui.lineEdit_surname.text()
            name = self.ui.lineEdit_name.text()

            # Άνοιγμα διαλόγου για επιλογή αρχείου
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Αποθήκευση Excel",
                f"{afm}_{surname}_{name}",
                "Excel Files (*.xlsx)"
            )
            if not file_path:
                return

            wb = openpyxl.Workbook()

            def autofit_columns(ws):
                for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            cell_len = len(str(cell.value)) if cell.value is not None else 0
                            if cell_len > max_length:
                                max_length = cell_len
                        except Exception:
                            pass # nosec B110 - cell width calc, cosmetic only
                    ws.column_dimensions[col_letter].width = max_length + 2
            # 1) ΦΥΛΛΟ: ΓΕΝΙΚΑ ΣΤΟΙΧΕΙΑ (Main Window)
        
            ws_main = wb.active
            ws_main.title = "TA Αρχικής"
            
           
            # Τίτλοι στηλών
            ws_main.cell(row=1, column=1, value="ΑΦΜ")
            ws_main.cell(row=1, column=2, value="Επώνυμο")
            ws_main.cell(row=1, column=3, value="Όνομα")
            
            # Παραλείπουμε τη στήλη "Διαγραφή" (col 9) στο export — δεν επηρεάζει
            # το import γιατί το _read_excel_file κάνει header-based matching.
            excel_col = 4
            for col in range(self.ta_page.model.columnCount()):
                if col == 9:
                    continue
                header = self.ta_page.model.headerData(col, Qt.Horizontal)
                ws_main.cell(row=1, column=excel_col, value=header)
                excel_col += 1
            # Δεδομένα
            for row in range(self.ta_page.model.rowCount()):
                # Στήλες 1-3: ΑΦΜ, Επώνυμο, Όνομα
                ws_main.cell(row=row+2, column=1, value=afm)
                ws_main.cell(row=row+2, column=2, value=surname)
                ws_main.cell(row=row+2, column=3, value=name)


            for row in range(self.ta_page.model.rowCount()):
                excel_col = 4
                for col in range(self.ta_page.model.columnCount()):
                    if col == 9:
                        continue
                    if col ==4:
                        value=self.ta_page.get_cell_strip(row,col)
                        if  value=="":
                            value="--Επιλέξτε"
                    elif col ==7:
                        value=self.ta_page.get_cell_strip(row,col)
                        crop_item = self.ta_page.model.item(row, 0)
                        crop = crop_item.text().strip()


                        if in_norm_set(crop, LOCK_AMPELI_NORM):

                            if value == "":
                                value = "--Επιλέξτε"
                        else:

                            value = ""


                    elif col in (2, 3, 5, 6, 8, 9, 10, 11, 12, 13, 14):
                        value = self.ta_page.get_cell_strip(row, col)
                        if not value:
                            value = None
                        else:
                            try:
                                value = float(value)
                            except ValueError:
                                value = None
                    else:
                        value = self.ta_page.get_cell_strip(row, col)

                    ws_main.cell(row=row+2, column=excel_col, value=value)
                    excel_col += 1
            
        
            # 3) ΦΥΛΛΟ: ΜΕΛΛΟΝΤΙΚΗ ΚΑΤΑΣΤΑΣΗ (mel_page)
        
            ws_mel = wb.create_sheet(title="ΤΑ Μελλοντικής")
            # Τίτλοι στηλών
            ws_mel.cell(row=1, column=1, value="ΑΦΜ")
            ws_mel.cell(row=1, column=2, value="Επώνυμο")
            ws_mel.cell(row=1, column=3, value="Όνομα")
            
            
            # Παραλείπουμε τη στήλη "Διαγραφή" (col 9) στο export — δεν επηρεάζει
            # το import γιατί το _read_excel_file κάνει header-based matching.
            excel_col = 4
            for col in range(self.mel_page.model.columnCount()):
                if col == 9:
                    continue
                header = self.mel_page.model.headerData(col, Qt.Horizontal)
                ws_mel.cell(row=1, column=excel_col, value=header)
                excel_col += 1
            # Δεδομένα
            for row in range(self.mel_page.model.rowCount()):
            # Στήλες 1-3: ΑΦΜ, Επώνυμο, Όνομα
                ws_mel.cell(row=row+2, column=1, value=afm)
                ws_mel.cell(row=row+2, column=2, value=surname)
                ws_mel.cell(row=row+2, column=3, value=name)


            for row in range(self.mel_page.model.rowCount()):
                excel_col = 4
                for col in range(self.mel_page.model.columnCount()):
                    if col == 9:
                        continue
                    if col ==4:
                        value=self.mel_page.get_cell_strip(row,col)
                        if  value=="":
                            value="--Επιλέξτε"
                    elif col ==7:
                        value=self.mel_page.get_cell_strip(row,col)
                        crop_item = self.mel_page.model.item(row, 0)
                        crop = crop_item.text().strip()


                        if in_norm_set(crop, LOCK_AMPELI_NORM):

                            if value == "":
                                value = "--Επιλέξτε"
                        else:

                            value = ""

                    elif col in (2, 3, 5, 6, 8, 9, 10, 11, 12, 13, 14):
                        value = self.mel_page.get_cell_strip(row, col)
                        if not value:
                            value = None
                        else:
                            try:
                                value = float(value)
                            except ValueError:
                                value = None
                    else:
                        value = self.mel_page.get_cell_strip(row, col)

                    ws_mel.cell(row=row+2, column=excel_col, value=value)
                    excel_col += 1
        
        
        
            # 4) ΦΥΛΛΟ: ΕΠΙΛΕΞΙΜΟΤΗΤΑ (epi_page) 
        
            ws_epi = wb.create_sheet(title="Επιλεξιμότητα")

            ws_epi.cell(row=1, column=1, value="Κριτήριο Επιλεξιμότητας")
            ws_epi.cell(row=1, column=2, value="Απάντηση")
            ws_epi.cell(row=1, column=3, value="Αποτέλεσμα")

            ws_epi.cell(row=2, column=1, value=self.epi_page.ui.label_2.text())
            ws_epi.cell(row=2, column=2, value=self.epi_page.ui.lineEdit_9.text())
            ws_epi.cell(row=2, column=3, value=self.epi_page.ui.lineEdit.text())

            ws_epi.cell(row=3, column=1, value=self.epi_page.ui.label_3.text())
            ws_epi.cell(row=3, column=2, value=self.epi_page.ui.comboBox_2.currentText())
            ws_epi.cell(row=3, column=3, value=self.epi_page.ui.lineEdit_2.text())

            ws_epi.cell(row=4, column=1, value=self.epi_page.ui.label_4.text())
            ws_epi.cell(row=4, column=2, value=self.epi_page.ui.comboBox_3.currentText())
            ws_epi.cell(row=4, column=3, value=self.epi_page.ui.lineEdit_3.text())

            ws_epi.cell(row=5, column=1, value=self.epi_page.ui.label_5.text())
            ws_epi.cell(row=5, column=2, value=self.epi_page.ui.comboBox_4.currentText())
            ws_epi.cell(row=5, column=3, value=self.epi_page.ui.lineEdit_4.text())

            ws_epi.cell(row=6, column=1, value=self.epi_page.ui.label_6.text())
            ws_epi.cell(row=6, column=2, value=self.epi_page.ui.comboBox_5.currentText())
            ws_epi.cell(row=6, column=3, value=self.epi_page.ui.lineEdit_5.text())

            ws_epi.cell(row=7, column=1, value=self.epi_page.ui.label_7.text())
            ws_epi.cell(row=7, column=2, value=self.epi_page.ui.comboBox_6.currentText())
            ws_epi.cell(row=7, column=3, value=self.epi_page.ui.lineEdit_6.text())

            ws_epi.cell(row=8, column=1, value=self.epi_page.ui.label_9.text())
            ws_epi.cell(row=8, column=2, value=self.epi_page.ui.comboBox.currentText())
            ws_epi.cell(row=8, column=3, value=self.epi_page.ui.lineEdit_10.text())

            ws_epi.cell(row=9, column=1, value=self.epi_page.ui.label_10.text())
            ws_epi.cell(row=9, column=2, value=self.epi_page.ui.comboBox_7.currentText())
            ws_epi.cell(row=9, column=3, value=self.epi_page.ui.lineEdit_11.text())

            ws_epi.cell(row=10, column=1, value=self.epi_page.ui.label_11.text())
            ws_epi.cell(row=10, column=2, value=self.epi_page.ui.comboBox_8.currentText())
            ws_epi.cell(row=10, column=3, value=self.epi_page.ui.lineEdit_12.text())

            ws_epi.cell(row=11, column=1, value=self.epi_page.ui.label_8.text())
            ws_epi.cell(row=11, column=2, value=to_float_or_none(self.epi_page.ui.lineEdit_7.text()))
            ws_epi.cell(row=11, column=3, value=self.epi_page.ui.lineEdit_8.text())

            ws_epi.cell(row=12, column=1, value=label_text(self.epi_page.ui.label_12))          
            ws_epi.cell(row=12, column=3, value=self.epi_page.ui.lineEdit_13.text())
            
            #5) φύλλο μοριοδότησης
            ws_moria = wb.create_sheet(title="Μοριοδότηση")

            ws_moria.cell(row=1, column=2, value=label_text(self.moria_page.ui.label_21))
            ws_moria.cell(row=1, column=3, value=to_float_or_none(self.moria_page.ui.lineEdit_budget.text()))

            ws_moria.cell(row=2, column=1, value="Κριτήρια Μοριοδότησης")
            ws_moria.cell(row=2, column=2, value="Απάντηση")
            ws_moria.cell(row=2, column=3, value="Μόρια")

            ws_moria.cell(row=3, column=1, value=self.moria_page.ui.label_3.text())
            ws_moria.cell(row=3, column=2, value=self.moria_page.ui.comboBox_1_1.currentText())
            ws_moria.cell(row=3, column=3, value=to_float_or_none(self.moria_page.ui.lineEdit_1_1_moria.text()))

            ws_moria.cell(row=4, column=1, value=self.moria_page.ui.label_4.text())
            ws_moria.cell(row=4, column=2, value=self.moria_page.ui.lineEdit_1_2.text())
            ws_moria.cell(row=4, column=3, value=to_float_or_none(self.moria_page.ui.lineEdit_1_2_moria.text()))

            ws_moria.cell(row=5, column=1, value=self.moria_page.ui.label_5.text())
            ws_moria.cell(row=5, column=2, value=to_float_or_none(self.moria_page.ui.lineEdit_2_1.text()))
            ws_moria.cell(row=5, column=3, value=to_float_or_none(self.moria_page.ui.lineEdit_2_1_moria.text()))

            ws_moria.cell(row=6, column=1, value=self.moria_page.ui.label_6.text())
            ws_moria.cell(row=6, column=2, value=to_float_or_none(self.moria_page.ui.lineEdit_2_2.text()))
            ws_moria.cell(row=6, column=3, value=to_float_or_none(self.moria_page.ui.lineEdit_2_2_moria.text()))

            ws_moria.cell(row=7, column=1, value=self.moria_page.ui.label_7.text())
            ws_moria.cell(row=7, column=2, value=to_float_or_none(self.moria_page.ui.lineEdit_3_1_1.text()))
            ws_moria.cell(row=7, column=3, value=to_float_or_none(self.moria_page.ui.lineEdit_3_1_1_moria.text()))

            ws_moria.cell(row=8, column=1, value=self.moria_page.ui.label_8.text())
            ws_moria.cell(row=8, column=2, value=self.moria_page.ui.comboBox_3_1_2.currentText())
            ws_moria.cell(row=8, column=3, value=to_float_or_none(self.moria_page.ui.lineEdit_3_1_2_moria.text()))

            ws_moria.cell(row=9, column=1, value=self.moria_page.ui.label_9.text())
            ws_moria.cell(row=9, column=2, value=self.moria_page.ui.comboBox_3_1_3.currentText())
            ws_moria.cell(row=9, column=3, value=to_float_or_none(self.moria_page.ui.lineEdit_3_1_3_moria.text()))

            ws_moria.cell(row=10, column=1, value=self.moria_page.ui.label_10.text())
            ws_moria.cell(row=10, column=2, value=self.moria_page.ui.comboBox_3_1_4.currentText())
            ws_moria.cell(row=10, column=3, value=to_float_or_none(self.moria_page.ui.lineEdit_3_1_4_moria.text()))

            ws_moria.cell(row=11, column=1, value=self.moria_page.ui.label_12.text())
            ws_moria.cell(row=11, column=2, value=self.moria_page.ui.comboBox_3_2.currentText())
            ws_moria.cell(row=11, column=3, value=to_float_or_none(self.moria_page.ui.lineEdit_3_2_moria.text()))

            ws_moria.cell(row=12, column=1, value=self.moria_page.ui.label_13.text())
            ws_moria.cell(row=12, column=2, value=self.moria_page.ui.comboBox_3_3.currentText())
            ws_moria.cell(row=12, column=3, value=to_float_or_none(self.moria_page.ui.lineEdit_3_3_moria.text()))

            ws_moria.cell(row=13, column=1, value=self.moria_page.ui.label_14.text())
            ws_moria.cell(row=13, column=2, value=self.moria_page.ui.comboBox_3_4.currentText())
            ws_moria.cell(row=13, column=3, value=to_float_or_none(self.moria_page.ui.lineEdit_3_4_moria.text()))

            ws_moria.cell(row=14, column=1, value=self.moria_page.ui.label_15.text())
            ws_moria.cell(row=14, column=2, value=self.moria_page.ui.comboBox_3_5.currentText())
            ws_moria.cell(row=14, column=3, value=to_float_or_none(self.moria_page.ui.lineEdit_3_5_moria.text()))

            ws_moria.cell(row=15, column=1, value=self.moria_page.ui.label_16.text())
            ws_moria.cell(row=15, column=2, value=to_float_or_none(self.moria_page.ui.lineEdit_4_1.text()))
            ws_moria.cell(row=15, column=3, value=to_float_or_none(self.moria_page.ui.lineEdit_4_1_moria.text()))

            ws_moria.cell(row=16, column=1, value=self.moria_page.ui.label_17.text())
            ws_moria.cell(row=16, column=2, value=to_float_or_none(self.moria_page.ui.lineEdit_5_1.text()))
            ws_moria.cell(row=16, column=3, value=to_float_or_none(self.moria_page.ui.lineEdit_5_1_moria.text()))

            ws_moria.cell(row=17, column=1, value=self.moria_page.ui.label_18.text())
            ws_moria.cell(row=17, column=2, value=self.moria_page.ui.comboBox_6_1.currentText())
            ws_moria.cell(row=17, column=3, value=to_float_or_none(self.moria_page.ui.lineEdit_6_1_moria.text()))

            ws_moria.cell(row=18, column=1, value=self.moria_page.ui.label_19.text())
            ws_moria.cell(row=18, column=2, value=self.moria_page.ui.comboBox_7_1.currentText())
            ws_moria.cell(row=18, column=3, value=to_float_or_none(self.moria_page.ui.lineEdit_7_1_moria.text()))

            ws_moria.cell(row=19, column=2, value=label_text(self.moria_page.ui.label_20))
            ws_moria.cell(row=19, column=3, value=to_float_or_none(self.moria_page.ui.lineEdit_moria.text()))

            ws_moria.cell(row=20, column=2, value=label_text(self.moria_page.ui.label_2))
            ws_moria.cell(row=20, column=3, value=self.moria_page.ui.lineEdit_epileximos.text())


            # Bold headers
            for cell in ws_main[1]:
                cell.font = Font(bold=True)
            for cell in ws_mel[1]:
                cell.font = Font(bold=True)
            for cell in ws_epi[1]:
                cell.font = Font(bold=True)
            for cell in ws_moria[2]:
                cell.font = Font(bold=True)

            # Autofit όλα τα φύλλα
            for ws in [ws_main, ws_mel, ws_epi, ws_moria]:
                autofit_columns(ws)

            # Αποθήκευση
            wb.save(file_path)
            show_temp_message_main(self, "Η εξαγωγή ολοκληρώθηκε!", bg_color="#4ABD56")
        except PermissionError:
            QMessageBox.critical(
                self,
                "Σφάλμα Δικαιωμάτων",
                "Δεν υπάρχουν δικαιώματα εγγραφής στον επιλεγμένο φάκελο.\n\n"
                "Επιλέξτε διαφορετική τοποθεσία."
            )
    
        except Exception as e:
            QMessageBox.critical(
                self,
                "Σφάλμα Εξαγωγής",
                f"Η εξαγωγή απέτυχε.\n\n"
                f"{str(e)}"
            )
    

    def _save_table_data(self, afm, page, scenario):
        """Βοηθητική μέθοδος για αποθήκευση πινάκωνν αρχικής-μελλοντικής ΤΑ"""
        
        if page.model.rowCount() == 0:
            # Αν δεν υπάρχουν γραμμές, διάγραψε παλιές εγγραφές
            save_scenario_data(afm, scenario, [])
            return
        
        raw_rows = page.get_all_rows_data()
        formatted_data = []
        
        for row in raw_rows:
            clean_row = (
                afm, scenario, 
                row[0], row[1], row[2], row[3], row[4], 
                row[5], row[6], row[7], row[8], row[10], 
                row[11], row[12], row[13], row[14]
            )
            formatted_data.append(clean_row)
        
        save_scenario_data(afm, scenario, formatted_data)

    
    def handle_save(self):
            """Def για το κουμπί αποθήκευσης"""
            try:
                afm = self.ui.lineEdit_afm.text().strip()
                if len(afm) != 9:
                    show_temp_message_main(self, "Ο ΑΦΜ πρέπει να είναι 9 ψηφία!")
                    return
                
                # Αποθήκευση/Ενημέρωση βασικών στοιχείων στη SQLite
                name = self.ui.lineEdit_name.text().strip()
                surname = self.ui.lineEdit_surname.text().strip()
                reg = self.ui.combo_periferia.currentText()
                save_producer_basics(afm, name, surname, reg)

                # Αποθήκευση Επιλεξιμότητας
                save_eligibility_data(afm, self.epi_page.get_eligibility_data())

                # Αποθήκευση Μοριοδότησης
                save_moria_data(afm, self.moria_page.get_moria_data())
                # Αποθήκευση ΟΛΩΝ των πινάκων
                self._save_table_data(afm, self.ta_page, 'initial')
                self._save_table_data(afm, self.mel_page, 'future')


                # Ανανέωση μόνο της γραμμής του τρέχοντος ΑΦΜ (αποφυγή πλήρους reload 4500+ γραμμών)
                self.arx_page.upsert_producer_row(afm)

                # Reset flag
                self.has_unsaved_changes = False
            
                show_temp_message_main(self, "Επιτυχής αποθήκευση!", bg_color="#4CAF50")
            except PermissionError:
                 QMessageBox.critical(
            self,
            "Σφάλμα Δικαιωμάτων",
            "Δεν υπάρχουν δικαιώματα εγγραφής στη βάση δεδομένων.\n\n"
            
        )
            except Exception as e:
                QMessageBox.critical(
                    self,
                    "Σφάλμα Αποθήκευσης",
                    f"Η αποθήκευση απέτυχε.\n\n"
                    f"{str(e)}\n\n"
                    
                )


    def handle_search(self):
        """Def για το κουμπί της φόρτωσης"""
        try:       
            afm = self.ui.lineEdit_afm.text()
            if len(afm) != 9: 
                show_temp_message_main(self, "Ο ΑΦΜ πρέπει να είναι 9 ψηφία!")
                return
            
            # Block signals
            self._block_all_signals(True)
            
            # Καθαρισμός
            self.ta_page.load_rows_from_data([])
            self.mel_page.load_rows_from_data([])
            self.epi_page.clear_page_epi()
            self.moria_page.clear_page_moria() 
            
            self.ui.lineEdit_name.clear()
            self.ui.lineEdit_surname.clear()
            self.ui.combo_periferia.setCurrentIndex(0)

            # Αναζήτηση παραγωγού
            producer = fetch_producer(afm)
            
            if producer:
                self.ui.lineEdit_name.setText(producer[0])
                self.ui.lineEdit_surname.setText(producer[1])
                
                
                region = producer[2]
                
                
                self.ui.combo_periferia.setCurrentText(region)
                
                
                
                self.ta_page.selected_periferia = region
                self.mel_page.selected_periferia = region
                self.epi_page.selected_periferia=region
                self.moria_page.selected_periferia=region
                

                # Φόρτωση Επιλεξιμότητας
                elig_data = fetch_eligibility(afm)
                if elig_data: 
                    self.epi_page.set_eligibility_data(elig_data)
                    self.epi_page.recalculate_all_results()

                #Φόρτωσης μοριοδότησης
                moria_data=fetch_moria(afm)
                if moria_data:
                    self.moria_page.set_moria_data(moria_data)
                    self.moria_page.recalculate_all_results_moria()
                
                # Φόρτωση δεδομένων πινάκων
                initial_rows = fetch_entries(afm, 'initial')
                formatted_initial = [[r[3], r[4], r[5], r[6], r[7], r[8], r[9], r[10], r[11], "", r[12], r[13], r[14], r[15], r[16]] for r in initial_rows]
                
                
                
                self.ta_page.load_rows_from_data(formatted_initial)
                
                missing_data = self._build_missing_check_data_from_db_rows(afm, initial_rows)
               
                missing_msg = self._check_missing_categories(missing_data, self._ta_categories_cache)

                self._pending_ta_warning_message = missing_msg

                # Αν ο χρήστης είναι ήδη στην TA σελίδα, δείξ’ το άμεσα
                self._show_pending_ta_warning()
                                
                future_rows = fetch_entries(afm, 'future')
                formatted_future = [[r[3], r[4], r[5], r[6], r[7], r[8], r[9], r[10], r[11], "", r[12], r[13], r[14], r[15], r[16]] for r in future_rows]
                self.mel_page.load_rows_from_data(formatted_future)


                missing_data_mel = self._build_missing_check_data_from_db_rows(afm, future_rows)
                missing_msg_mel = self._check_missing_categories(missing_data_mel, self._ta_categories_cache)
                self._pending_mel_warning_message = missing_msg_mel
                self._show_pending_mel_warning()

                self._current_loaded_afm = afm
                
                show_temp_message_main(self, "Τα δεδομένα φορτώθηκαν!", bg_color="#4CAF50")
            else:
                show_temp_message_main(self, "Το ΑΦΜ δεν βρέθηκε στη βάση.\nΠροχωρήστε σε Αποθήκευση για να πραγματοποιηθεί Νέα Εγγραφή του παραγωγού")
            
            # Enable τα κουμπιά (και για υπάρχον και για νέο AFM)
            self.search_or_edit_performed = True
            self._lock_actions_only(True)
            self.ui.savebtn.setEnabled(True)
            self.ui.exportbtn.setEnabled(True)
            self.ui.importbtn.setEnabled(False)
            self.has_unsaved_changes = False
            self._block_all_signals(False)
        
        except Exception as e:
            self._block_all_signals(False)
            QMessageBox.critical(self, "Σφάλμα Αναζήτησης", f"Η αναζήτηση απέτυχε.\n\n{str(e)}")
   

    def clear_ui(self):
        """Καθαρίζει όλα τα πεδία και τους πίνακες της οθόνης."""
        
        # Καθαρισμός κειμένων
        self.ui.lineEdit_afm.clear()
        self.ui.lineEdit_name.clear()
        self.ui.lineEdit_surname.clear()
        self.ui.combo_periferia.setCurrentIndex(0)
    
        # Καθαρισμός πινάκων,  epi-page και μοριοδότησης
        self.ta_page.load_rows_from_data([])
        self.mel_page.load_rows_from_data([])
        self.epi_page.clear_page_epi()
        self.moria_page.clear_page_moria() 

        
        self.ta_page.selected_periferia = DEFAULT
        self.mel_page.selected_periferia = DEFAULT
        self.epi_page.selected_periferia = DEFAULT
        self.moria_page.selected_periferia = DEFAULT

        # disable τα κουμπιά (και για υπάρχον και για νέο AFM)
        self.search_or_edit_performed = False
        self.ui.savebtn.setEnabled(False)
        self.ui.exportbtn.setEnabled(False)

        # Καθαρισμός deferred warnings — ανήκουν στο προηγούμενο AFM
        self._pending_ta_warning_message = None
        self._pending_mel_warning_message = None


    def clear_ui_not_afm(self):
        """Helper: Καθαρισμός πεδίων και πινάκων (χωρίς AFM)"""

        self._block_all_signals(True)
        self.ui.lineEdit_name.clear()
        self.ui.lineEdit_surname.clear()
        self.ui.combo_periferia.setCurrentIndex(0)

        self.ta_page.selected_periferia = DEFAULT
        self.mel_page.selected_periferia = DEFAULT
        self.epi_page.selected_periferia = DEFAULT
        self.moria_page.selected_periferia = DEFAULT

        self.ta_page.load_rows_from_data([])
        self.mel_page.load_rows_from_data([])
        self.epi_page.clear_page_epi()
        self.moria_page.clear_page_moria()

        # Καθαρισμός deferred warnings — το προηγούμενο AFM "έχασε" την ευκαιρία
        self._pending_ta_warning_message = None
        self._pending_mel_warning_message = None

        self._block_all_signals(False)
        
     
    
    def _setup_minimum_sizes(self):
        """Minimum window size για Windows 7+"""
        
        self.setMinimumSize(900, 650)


# EXCEL & CSV IMPORT


    def import_data_from_file(self):
        """Import CSV ή Excel αρχείου"""
        try:
            # 1. Επιλογή αρχείου
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Επιλογή Αρχείου για Import",
                "",
                "Excel/CSV Files (*.xlsx *.csv);;Excel Files (*.xlsx);;CSV Files (*.csv)"
            )

            if not file_path:
                return
            
            # Progress για ανάγνωση
            read_progress = QProgressDialog("Ανάγνωση αρχείου...", None, 0, 0, self)
            read_progress.setWindowModality(Qt.WindowModal)
            read_progress.setMinimumDuration(500)
            read_progress.setCancelButton(None)
            read_progress.setValue(0)

            try:
                # 2. Ανάγνωση δεδομένων
                if file_path.endswith('.csv'):
                    import_data, skipped_afms = self._read_csv_file(file_path)
                elif file_path.endswith('.xlsx'):
                    import_data, skipped_afms = self._read_excel_file(file_path)
                else:
                    QMessageBox.warning(
                        self,
                        "Μη Υποστηριζόμενος Τύπος",
                        "Υποστηρίζονται μόνο αρχεία .csv και .xlsx"
                    )
                    return
            finally:
                read_progress.close()

            total_skipped =len(skipped_afms)
        
            if total_skipped > 0:
                skipped_msg = "Παραλείφθηκαν κατά την ανάγνωση:\n\n"
                
                
                skipped_msg += f"[X] Μη έγκυροι ΑΦΜ: {len(set(skipped_afms))}\n"
                
                
                
                skipped_msg += f"\n[OK] Έγκυρα ΑΦΜ αρχείου προς εισαγωγή: {len(import_data)}"
                
                QMessageBox.information(self, "Πληροφορίες Ανάγνωσης", skipped_msg)
            
            if not import_data:
                QMessageBox.information(
                    self,
                    "Σφάλμα Δεδομένων Αρχείου",
                    "Το αρχείο δεν περιέχει έγκυρα δεδομένα."
                )
                return
            
            # 3. Διαχωρισμός σε νέα και υπάρχοντα AFM
            all_producers = fetch_all_producers()
            existing_afms = {str(p[0]) for p in all_producers}
            
            new_data = []
            conflicting_data = []
            
            for data in import_data:
                if data['afm'] in existing_afms:
                    conflicting_data.append(data)
                else:
                    new_data.append(data)
            
            #  4. Αν υπάρχουν conflicts, εμφάνιση dialog
            replace_afms = set()  # AFM που θα αντικατασταθούν
            
            if conflicting_data:
                dialog = ImportConflictDialog(conflicting_data, self)
                
                if dialog.exec_() == QDialog.Accepted:
                    replace_afms = dialog.get_selected_afms()  
                else:
                    # Ακύρωση - import μόνο νέα
                    if not new_data:
                        return
            
            
            # 5. Import δεδομένων
            self._import_data(new_data, conflicting_data, replace_afms)

            #κλείδωμα Lock actions
            # self._lock_actions_only(False)
            # self.ui.savebtn.setEnabled(False)
            # self.ui.exportbtn.setEnabled(False)

            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Σφάλμα Εισαγωγής",
                f"Η εισαγωγή απέτυχε.\n\n{str(e)}"
            )
        

    
    def _import_data(self, new_data, conflicting_data, replace_afms):
        """Import με error recovery"""
        
        total_afms = len(new_data) + len(replace_afms)
        
        if total_afms == 0:
            QMessageBox.information(self, "Πληροφορία", "Δεν επιλέχθηκε κανένα ΑΦΜ.")
            return
        
        # Lock UI
        self._lock_ui_during_import(True)

        # Ένα progress bar, μόνο για την εισαγωγή των ΑΦΜ στη βάση.
        # Το refresh του πίνακα μετά το import είναι πλέον γρήγορο (pre-aggregated
        # LEFT JOIN + index) οπότε δεν χρειάζεται δικό του progress — γίνεται
        # σιωπηλά με το UI ακόμα κλειδωμένο (wait cursor).
        progress = QProgressDialog(
            "Εισαγωγή δεδομένων...",
            None,
            0,
            total_afms,
            self
        )
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setAutoClose(False)
        progress.setAutoReset(False)
        progress.setWindowTitle("Εισαγωγή δεδομένων")

        progress.setStyleSheet("""
            QProgressDialog {
                min-width: 400px;
                min-height: 150px;
                background-color: #F2F7FA;
            }
            QLabel {
                font-size: 12pt;
                padding: 10px;
            }
        """)
        progress.setValue(0)
        QApplication.processEvents()

        success_count = 0
        replace_count=0
        failed_afms = []
        new_count=0


        try:
            # Prepare data
            all_data_to_import = []
            all_data_to_import.extend(new_data)

            for data in conflicting_data:
                if data['afm'] in replace_afms:
                    all_data_to_import.append({**data, '_replace': True})

            # DB import με progress callback
            result = import_producers_batch_transaction(
                all_data_to_import,
                progress_callback=lambda current, total: (
                    progress.setValue(current),
                    progress.setLabelText(
                        f"Εισαγωγή ΑΦΜ...\n\n({current}/{total})"
                    ),
                    QApplication.processEvents()
                )
            )

            #  Extract results
            success_count = result['total_success']
            failed_afms = result['failed']
            replace_count = result['replace']
            new_count = success_count - replace_count

        except Exception as e:
            #  ΜΟΝΟ ΓΙΑ ΣΟΒΑΡΑ ΣΦΑΛΜΑΤΑ (DB connection, disk full, etc.)
            QMessageBox.critical(
                self,
                "Σφάλμα Import",
                f"Το import απέτυχε λόγω σοβαρού προβλήματος.\n\n{str(e)}"
            )
        finally:
            progress.close()

        # Σιωπηλό refresh του πίνακα — το UI παραμένει κλειδωμένο με wait cursor
        # μέχρι να ολοκληρωθεί (συνήθως <1s χάρη στο LEFT JOIN + index).
        try:
            if hasattr(self, 'arx_page'):
                # Καθαρισμός φίλτρου πριν το refresh — ο χρήστης βλέπει όλη τη νέα
                # λίστα, και αποφεύγουμε edge cases με stale filter mapping.
                self.arx_page.clear_filter()
                self.arx_page.load_producers()
        finally:
            self._lock_ui_during_import(False)

        #  Show detailed summary
        self._show_import_summary(success_count, failed_afms, new_count, replace_count)


    def _show_import_summary(self, success_count, failed_afms, new_count, replace_count):
        """Εμφάνιση detailed import summary"""
        #κλείδωμα Lock actions
        self._lock_actions_only(False)
        self.ui.savebtn.setEnabled(False)
        self.ui.exportbtn.setEnabled(False)
        
        error_count = len(failed_afms)
        
        # Complete success
        if error_count == 0 and success_count > 0 and replace_count==0:
            msg = (
                f"Επιτυχής εισαγωγή των ΑΦΜ!\n\n"
                f"Σύνολο νέων ΑΦΜ: {success_count} ΑΦΜ\n\n"
                f"(*) Για κάθε νέο ΑΦΜ επιλέξτε Περιφέρεια\n"
                f"για να υπολογιστεί η Τυπική Απόδοση."
            )
            QMessageBox.information(self, "Επιτυχής Εισαγωγή", msg)
    
        elif error_count == 0  and new_count==0 and replace_count>0:
            msg = (
                f"Επιτυχής εισαγωγή των ΑΦΜ!\n\n"
                f"Σύνολο ΑΦΜ που αντικαταστάθηκαν : {replace_count} ΑΦΜ\n\n"
                f"(*) Προσοχή : Για τους ΑΦΜ που αντικαταστάθηκαν ενημερώθηκε  ο πίνακας της Αρχικής ΤΑ.\n"
                f"-Θα πρέπει να ενημερώσετε τον πίνακα της Μελλοντικής ΤΑ εφόσον θέλετε να περαστούν και εκεί οι αλλαγές.\n"
                f"-Θα πρέπει για κάθε ΑΦΜ που αντικαταστάθηκε να πατήσετε Αποθήκευση για να ενημερωθεί ο Συγκεντρωτικός Πίνακας Εγγραφών"
            )
            QMessageBox.information(self, "Επιτυχής Εισαγωγή", msg)
        
        elif error_count == 0 and new_count > 0 and replace_count>0:
            msg = (
                f"Επιτυχής εισαγωγή των ΑΦΜ!\n\n"
                f"Σύνολο νέων ΑΦΜ: {new_count} ΑΦΜ\n\n"
                f"Σύνολο ΑΦΜ που αντικαταστάθηκαν : {replace_count} ΑΦΜ\n\n"
                f"(*) Για κάθε νέο ΑΦΜ επιλέξτε Περιφέρεια για να υπολογιστεί η Τυπική Απόδοση.\n"
                f"(**) Προσοχή : Για τους ΑΦΜ που αντικαταστάθηκαν ενημερώθηκε  ο πίνακας της Αρχικής ΤΑ.\n"
                f"-Θα πρέπει να ενημερώσετε τον πίνακα της Μελλοντικής ΤΑ εφόσον θέλετε να περαστούν και εκεί οι αλλαγές.\n"
                f"-Θα πρέπει για κάθε ΑΦΜ που αντικαταστάθηκε να πατήσετε Αποθήκευση για να ενημερωθεί ο Συγκεντρωτικός Πίνακας Εγγραφών"
            )
            QMessageBox.information(self, "Επιτυχής Εισαγωγή", msg)

            
       
        
        # Partial success
        elif success_count > 0 and error_count > 0:
            failed_list = "\n".join(
                f"  • {afm}: {error}" 
                for afm, error in failed_afms[:10]
            )
            
            if len(failed_afms) > 10:
                failed_list += f"\n  ... και {len(failed_afms) - 10} ακόμα"
            
            msg = (
                f"Μερική Εισαγωγή\n\n"
                f"Επιτυχής εισαγωγή: {success_count} ΑΦΜ\n"
                f"Αποτυχίες: {error_count} ΑΦΜ\n\n"
                f"Προβληματικά ΑΦΜ:\n{failed_list}\n\n"
                f"(*) Ελέγξτε τα δεδομένα και δοκιμάστε ξανά."
            )
            
            QMessageBox.warning(self, "Μερική Εισαγωγή", msg)
        
        # Complete failure
        else:
            msg = f"[X] Κανένα ΑΦΜ δεν εισήχθη.\n\nΑποτυχίες: {error_count}"
            QMessageBox.critical(self, "Αποτυχία Import", msg)

        

    
    def _lock_ui_during_import(self, locked):
        if hasattr(self.ui, 'centralwidget'):
            self.ui.centralwidget.setEnabled(not locked)

        # QActions (menu) για αλλαγή σελίδας
        for attr in ('arxiki', 'ta', 'mellontiki', 'epileximotita', 'moria'):
            if hasattr(self.ui, attr):
                getattr(self.ui, attr).setEnabled(not locked)

        # Ρητή απενεργοποίηση του πίνακα της αρχικής + των ✎/× delegates
        # Το viewport().setEnabled(False) είναι κρίσιμο: τα painted delegates
        # λαμβάνουν clicks μέσω editorEvent στο viewport, όχι στο ίδιο το tableView.
        if hasattr(self, 'arx_page'):
            tv = self.arx_page.ui.tableView
            tv.setEnabled(not locked)
            tv.viewport().setEnabled(not locked)
            if hasattr(self.arx_page.ui, 'lineEdit_search'):
                self.arx_page.ui.lineEdit_search.setEnabled(not locked)
            if hasattr(self.arx_page.ui, 'pushButton'):
                self.arx_page.ui.pushButton.setEnabled(not locked)

        # Top-bar κουμπιά & ΑΦΜ input
        for attr in ('importbtn', 'exportbtn', 'savebtn', 'searchbtn', 'lineEdit_afm'):
            if hasattr(self.ui, attr):
                getattr(self.ui, attr).setEnabled(not locked)

        if locked:
            QApplication.setOverrideCursor(Qt.WaitCursor)
        else:
            QApplication.restoreOverrideCursor()

    # Σταθερές validation για import (Excel/CSV)
    _ALLOWED_CERTS = {"--Επιλέξτε", "Συμβατικά", "Βιολογικά", "Ολοκληρωμένη", "ΠΟΠ/ΠΓΕ"}
    _ALLOWED_VINE_XLSX = {"--Επιλέξτε", "Ναι", "Όχι"}
    _ALLOWED_VINE_CSV = {"1", "0"}
    _MAX_QUANTITY = 9999999.99
    _MAX_INT_7 = 9999999
    _MAX_TEXT_LEN = 200

    def _canonicalize_entry_row(self, entry_row):
        """Αν η (cat, desc) ταιριάζει σε normalized επίπεδο με ζευγάρι του ta.xlsx,
        αντικαθιστά με τα κανονικά Ελληνικά. Mutates in place."""
        cat = (entry_row.get('category_osde') or '').strip()
        desc = (entry_row.get('description') or '').strip()

        if not cat:
            return

        norm_cat = norm(cat)

        # 1) Δοκίμασε πρώτα ολόκληρο το ζευγάρι (cat, desc)
        if desc:
            norm_desc = norm(desc)
            canonical = self._canonical_pair_by_norm.get((norm_cat, norm_desc))
            if canonical is not None:
                entry_row['category_osde'] = canonical[0]
                entry_row['description'] = canonical[1]
                return

        # 2) Αν δεν βρέθηκε pair, δοκίμασε μόνο cat
        canonical_cat = self._canonical_cat_by_norm.get(norm_cat)
        if canonical_cat is not None:
            entry_row['category_osde'] = canonical_cat

    def _validate_import_row(self, entry_row, source_format, raw_vine_csv=None):
        """Σηκώνει ValueError('Μη έγκυρα δεδομένα') στο πρώτο σφάλμα."""

        # Κατηγορία ΟΣΔΕ — αν δεν είναι κενή & δεν είναι έγκυρη, max 200 χαρ
        cat = (entry_row.get('category_osde') or '').strip()
        if cat and norm(cat) not in self._valid_cats_cache and len(cat) > self._MAX_TEXT_LEN:
            raise ValueError("Μη έγκυρα δεδομένα")

        # Περιγραφή Είδους — όμοια
        desc = (entry_row.get('description') or '').strip()
        if desc and norm(desc) not in self._valid_descs_cache and len(desc) > self._MAX_TEXT_LEN:
            raise ValueError("Μη έγκυρα δεδομένα")

        # Έκταση/Αριθμός ζώων — float [0, 9999999.99] με max 2 δεκαδικά
        qty = entry_row.get('quantity', '')
        qty_str = str(qty).strip() if qty is not None else ''
        if qty_str:
            qty_normalized = qty_str.replace(',', '.')
            try:
                qty_f = float(qty_normalized)
            except (ValueError, TypeError):
                raise ValueError("Μη έγκυρα δεδομένα")
            if qty_f < 0 or qty_f > self._MAX_QUANTITY:
                raise ValueError("Μη έγκυρα δεδομένα")
            # max 2 δεκαδικά ψηφία
            if '.' in qty_normalized and len(qty_normalized.split('.', 1)[1]) > 2:
                raise ValueError("Μη έγκυρα δεδομένα")

        # Δένδρα >=4 / <4 ετών — int [0, 9999999], όχι δεκαδικά
        for field in ('trees_over_4', 'trees_under_4'):
            val = entry_row.get(field, '')
            val_str = str(val).strip() if val is not None else ''
            if val_str:
                try:
                    val_f = float(val_str.replace(',', '.'))
                except (ValueError, TypeError):
                    raise ValueError("Μη έγκυρα δεδομένα")
                # Όχι δεκαδικά (12.0 ok από Excel, 12.5 fail)
                if val_f != int(val_f):
                    raise ValueError("Μη έγκυρα δεδομένα")
                val_i = int(val_f)
                if val_i < 0 or val_i > self._MAX_INT_7:
                    raise ValueError("Μη έγκυρα δεδομένα")

        # Βιολογικά — μόνο Excel, ΥΠΟΧΡΕΩΤΙΚΟ
        if source_format == 'excel':
            cert = (entry_row.get('certification') or '').strip()
            if cert not in self._ALLOWED_CERTS:
                raise ValueError("Μη έγκυρα δεδομένα")

        # Αμπέλι — μόνο όταν Κατηγορία ∈ LOCK_AMPELI_NORM
        if cat and in_norm_set(cat, LOCK_AMPELI_NORM):
            if source_format == 'excel':
                vine = (entry_row.get('vine_over_3') or '').strip()
                if vine and vine not in self._ALLOWED_VINE_XLSX:
                    raise ValueError("Μη έγκυρα δεδομένα")
            else:  # csv
                v = str(raw_vine_csv or '').strip()
                if v and v not in self._ALLOWED_VINE_CSV:
                    raise ValueError("Μη έγκυρα δεδομένα")

    def _read_excel_file(self, file_path):
        """Διάβασμα Excel αρχείου - ΜΟΝΟ δεδομένα εισόδου"""
        import_data = []
        skipped_afms = []

        # Magic bytes check για XLSX (ZIP-based: ξεκινά με 'PK\x03\x04').
        # Αποτρέπει BadZipFile exceptions με αόριστο μήνυμα όταν το αρχείο
        # δεν είναι πραγματικά .xlsx (π.χ. .exe ή PDF με αλλαγμένη επέκταση).
        try:
            with open(file_path, 'rb') as f:
                magic = f.read(4)
            if magic[:2] != b'PK':
                raise ValueError(
                    "Το αρχείο δεν είναι έγκυρο Excel (.xlsx). "
                    "Πιθανόν να έχει αλλαχθεί η επέκταση ή να είναι κατεστραμμένο."
                )
        except OSError as e:
            raise Exception(f"Σφάλμα ανάγνωσης αρχείου: {e}")

        wb = None
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            if "TA Αρχικής" not in wb.sheetnames:
                raise ValueError("Δεν βρέθηκε το φύλλο 'TA Αρχικής' στο αρχείο Excel!")
            ws = wb["TA Αρχικής"]
            
            # Κανονικοποίηση whitespace: τα headers του UI περιέχουν '\n' (π.χ. "Δένδρα\n>=4 ετών")
            # για multi-line εμφάνιση — τα κάνουμε ενιαίο κενό για να ταιριάξουν με τα patterns παρακάτω.
            headers = [' '.join(str(cell.value).split()) if cell.value else '' for cell in ws[1]]
            
            # Mapping ΜΟΝΟ των δεδομένων εισόδου (ΟΧΙ calculated fields)
            col_map = {}
            for i, header in enumerate(headers):
                h_lower = header.lower()
                
                if 'αφμ' in h_lower or h_lower == 'afm':
                    col_map['afm'] = i
                elif 'επώνυμο' in h_lower or 'surname' in h_lower:
                    col_map['surname'] = i
                elif 'όνομα' in h_lower or 'name' in h_lower:
                    col_map['name'] = i
                elif 'κατηγορία οσδε' in h_lower or 'category' in h_lower:
                    col_map['category_osde'] = i
                elif 'περιγραφή' in h_lower or 'description' in h_lower:
                    col_map['description'] = i
                elif 'έκταση' in h_lower or 'αριθμός ζώων' in h_lower or 'quantity' in h_lower:
                    col_map['quantity'] = i
                elif 'βιολογικά' in h_lower or 'certification' in h_lower:
                    col_map['certification'] = i
                elif 'δένδρα>=4' in h_lower or 'δένδρα >=4' in h_lower or 'trees_over_4' in h_lower:
                    col_map['trees_over_4'] = i
                elif 'δένδρα<4' in h_lower or 'δένδρα <4' in h_lower or 'trees_under_4' in h_lower:
                    col_map['trees_under_4'] = i
                elif 'αμπέλι>3' in h_lower or 'αμπέλι >3' in h_lower or 'vine_over_3' in h_lower:
                    col_map['vine_over_3'] = i
            
            # Αυστηρό validation: όλες οι αναμενόμενες στήλες πρέπει να υπάρχουν,
            # αλλιώς ακυρώνουμε εξ ολοκλήρου την εισαγωγή για λόγους ασφάλειας
            # (αποφυγή σιωπηλά ελλιπών δεδομένων).
            REQUIRED_EXCEL_COLS = {
                'afm':           "ΑΦΜ",
                'name':          "Όνομα",
                'surname':       "Επώνυμο",
                'category_osde': "Κατηγορία ΟΣΔΕ",
                'description':   "Περιγραφή Είδους/Ποικιλίας/Ζώων",
                'quantity':      "Έκταση/Αριθμός ζώων",
                'certification': "Βιολογικά Ολοκλ/μένη ΠΟΠ/ΠΓΕ",
                'trees_over_4':  "Δένδρα >=4 ετών",
                'trees_under_4': "Δένδρα <4 ετών",
                'vine_over_3':   "Αμπέλι >3 ετών",
            }
            missing_cols = [label for key, label in REQUIRED_EXCEL_COLS.items() if key not in col_map]
            if missing_cols:
                raise ValueError(
                    "Λείπουν υποχρεωτικές στήλες από το αρχείο Excel")
                

            #  Helper για ασφαλή ανάγνωση
            def get_cell_value(row, col_name):
                """Επιστρέφει τιμή από row ή '' αν λείπει στήλη"""
                col_idx = col_map.get(col_name)
                if col_idx is None or col_idx >= len(row):
                    return ''
                value = row[col_idx]
                if value is None or value == '':
                    return ''
                return str(value).strip()
            
            # Group rows by AFM
            afm_groups = {}
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                afm = get_cell_value(row, 'afm')
                
                if not afm or len(afm) != 9:
                    skipped_afms.append(afm)
                    continue
                
                if afm not in afm_groups:
                    afm_groups[afm] = {
                        'afm': afm,
                        'name': get_cell_value(row, 'name'),
                        'surname': get_cell_value(row, 'surname'),
                        'region': '--Επιλέξτε',
                        'rows': []
                    }
                
                # Αποθήκευση ΜΟΝΟ δεδομένων εισόδου
                entry_row = {
                    'category_osde': get_cell_value(row, 'category_osde'),
                    'description': get_cell_value(row, 'description'),
                    'quantity': get_cell_value(row, 'quantity'),
                    'certification': get_cell_value(row, 'certification'),
                    'trees_over_4': get_cell_value(row, 'trees_over_4'),
                    'trees_under_4': get_cell_value(row, 'trees_under_4'),
                    'vine_over_3': get_cell_value(row, 'vine_over_3'),
                    
                    # Calculated fields: ΠΑΝΤΑ ΚΕΝΑ
                    'typical_output': '',
                    'output_per_choice': '',
                    'total_output': '',
                    'ta_productive': '',
                    'ta_plant': '',
                    'ta_animal': '',
                    'ta_bees': '',
                }

                self._canonicalize_entry_row(entry_row)
                self._validate_import_row(entry_row, 'excel')
                afm_groups[afm]['rows'].append(entry_row)
            
            import_data = list(afm_groups.values())

        except ValueError:
            raise
        except Exception as e:
            raise Exception(f"Σφάλμα ανάγνωσης Excel: {str(e)}")

        finally:
            if wb is not None:
                wb.close()

        return import_data, skipped_afms
    
    
    

    def _read_csv_file(self, file_path):
        """Διάβασμα CSV αρχείου"""
       
        
        encodings = ['utf-8-sig', 'utf-8', 'cp1253', 'iso-8859-7', 'latin1']

        skipped_afms=[]
        
        
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding, newline='') as file:
                    sample = file.read(1024)
                    file.seek(0)
                    
                    delimiter = ',' if sample.count(',') > sample.count(';') else ';'
                    
                    csv_reader = csv.DictReader(file, delimiter=delimiter)

                    # Αυστηρό validation headers — ίδια λογική με Excel.
                    # Αν λείπει έστω και μία στήλη, ακυρώνουμε εντελώς το import.
                    raw_fieldnames = csv_reader.fieldnames or []
                    fieldnames_upper = {(fn or '').strip().upper() for fn in raw_fieldnames}
                    REQUIRED_CSV_COLS = [
                        ('ΑΦΜ',                  ('ΑΦΜ', 'AFM')),
                        ('Όνομα',                ('ΟΝΟΜΑ', 'NAME')),
                        ('Επώνυμο',              ('ΕΠΩΝΥΜΟ', 'SURNAME')),
                        ('Κατηγορία',            ('ΚΑΤΗΓΟΡΙΑ',)),
                        ('Περιγραφή',            ('ΠΕΡΙΓΡΑΦΗ',)),
                        ('Έκταση/Αριθμός ζώων',  ('ΕΚΤΑΣΗ_ΖΩΑ',)),
                        ('Δένδρα >=4 ετών',      ('ΔΕΝΤΡΑ_ΑΝΩ_4_ΕΤΩΝ',)),
                        ('Δένδρα <4 ετών',       ('ΔΕΝΤΡΑ_ΚΑΤΩ_4_ΕΤΩΝ',)),
                        ('Αμπέλι >3 ετών',       ('ΑΜΠΕΛ_3_ΕΤΩΝ',)),
                    ]
                    missing_cols = [
                        label for label, aliases in REQUIRED_CSV_COLS
                        if not any(a in fieldnames_upper for a in aliases)
                    ]
                    if missing_cols:
                        raise ValueError(
                            "Λείπουν υποχρεωτικές στήλες από το αρχείο CSV"
                        )

                    # Group rows by AFM
                    afm_groups = {}
                    
                    for row in csv_reader:
                        # Normalize keys
                        row_normalized = {k.strip().upper(): v.strip() if v else '' for k, v in row.items()}

                        # Validation + check ΕΠΙΣΠΟΡΗ
                        epispora = row_normalized.get('ΕΠΙΣΠΟΡΗ', '')
                        epispora_norm = norm(epispora)
                        if epispora_norm not in ('', NORM_YES, norm('ΟΧΙ')):
                            raise ValueError("Μη έγκυρα δεδομένα")
                        if epispora_norm == NORM_YES:
                            continue  # Skip this row (επισπορά)

                        # Get AFM
                        afm = row_normalized.get('ΑΦΜ') or row_normalized.get('AFM', '')

                        if not afm or len(afm) != 9:
                            skipped_afms.append(afm)
                            continue

                        # First time seeing this AFM
                        if afm not in afm_groups:
                            afm_groups[afm] = {
                                'afm': afm,
                                'name': row_normalized.get('ΟΝΟΜΑ') or row_normalized.get('NAME', ''),
                                'surname': row_normalized.get('ΕΠΩΝΥΜΟ') or row_normalized.get('SURNAME', ''),
                                'region': '--Επιλέξτε',  # Always default
                                'rows': []
                            }
                        
                        #  Parse entry row
                        entry_row = self._parse_csv_entry_row(row_normalized)

                        if entry_row:
                            vine_raw_csv = (row_normalized.get('ΑΜΠΕΛ_3_ΕΤΩΝ') or '').strip()
                            self._canonicalize_entry_row(entry_row)
                            self._validate_import_row(entry_row, 'csv', raw_vine_csv=vine_raw_csv)
                            afm_groups[afm]['rows'].append(entry_row)
                    
                    #  Remove producers with no entries (all rows had ΕΠΙΣΠΟΡΗ = ΝΑΙ)
                    import_data = [data for data in afm_groups.values() if data['rows']]
                    
                    return import_data, skipped_afms
                
            except UnicodeDecodeError:
                continue
            except ValueError:
                # Header validation error — δομικό πρόβλημα του αρχείου,
                # δεν σχετίζεται με encoding. Διακοπή όλης της διαδικασίας.
                raise
            except Exception as e:
                print(f"Σφάλμα ανάγνωσης {encoding}: {e}")
                continue
        
        return [],  skipped_afms

    

    def _parse_csv_entry_row(self, row):
        """Parse μιας σειράς CSV σε entry row"""
        
        #  Quantity (ha/sum) - Handle comma as decimal separator
        quantity = row.get('ΕΚΤΑΣΗ_ΖΩΑ')
        if quantity:
            # Replace comma with dot if it's a decimal separator
            # Check if there's only one comma (decimal) vs multiple (CSV delimiter already handled)
            quantity = quantity.replace(',', '.')
        
        #  Vine over 3 mapping
        vine_raw = row.get('ΑΜΠΕΛ_3_ΕΤΩΝ')
        vine_over_3 = ''
        if vine_raw == '1':
            vine_over_3 = 'Ναι'
        elif vine_raw == '0':
            vine_over_3 = 'Όχι'
        # else: κενό

        entry_row = {
            'category_osde': row.get('ΚΑΤΗΓΟΡΙΑ', ''),
            'description': row.get('ΠΕΡΙΓΡΑΦΗ'),
            'typical_output': '',  # Always empty for CSV
            'quantity': quantity,
            'certification': '',  # Always empty for CSV
            'trees_over_4': row.get('ΔΕΝΤΡΑ_ΑΝΩ_4_ΕΤΩΝ', ''),
            'trees_under_4': row.get('ΔΕΝΤΡΑ_ΚΑΤΩ_4_ΕΤΩΝ', ''),
            'vine_over_3': vine_over_3,
            'output_per_choice': '',  # Always empty for CSV
            'total_output': '',  # Always empty for CSV
            'ta_productive': '',  # Always empty for CSV
            'ta_plant': '',  # Always empty for CSV
            'ta_animal': '',  # Always empty for CSV
            'ta_bees': '',  # Always empty for CSV
        }
        
        return entry_row
    
    
    

    def _check_missing_categories(self, data, ta_categories=None):
        """Έλεγχος αν καλλιέργειες/ποικιλίες υπάρχουν στο ta.xlsx
        και αν το αμπέλι έχει συμπληρωμένη τη στήλη vine_over_3"""

        if ta_categories is None:
            ta_categories = self._ta_categories_cache

        

        missing = []        # εγγραφές χωρίς έγκυρη κατηγορία/περιγραφή
        missing_vine = []   # εγγραφές αμπελιού χωρίς τιμή στο vine_over_3

        for row_num, entry in enumerate(data.get('rows', []), start=1):
            category = (entry.get('category_osde') or '').strip()
            description = (entry.get('description') or '').strip()

            # Πλήρως κενές γραμμές → αγνοούνται
            if not category and not description:
                continue

            # Λείπει είτε η Κατηγορία είτε η Περιγραφή → invalid pair
            if not category or not description:
                cat_disp = category or '--Επιλέξτε'
                desc_disp = description or '--Επιλέξτε'
                missing.append(f"Σειρά {row_num}: {cat_disp} / {desc_disp}")
            else:
                # Έλεγχος 1: ύπαρξη στο ta.xlsx
                key = (norm(category), norm(description))
                if key not in ta_categories:
                    missing.append(f"Σειρά {row_num}: {category} / {description}")

            # Έλεγχος 2: αμπέλι χωρίς vine_over_3 (μόνο αν υπάρχει κατηγορία)
            if category and in_norm_set(category, LOCK_AMPELI_NORM):
                vine = (entry.get('vine_over_3') or '').strip()
                if vine not in ('Ναι', 'Όχι'):
                    missing_vine.append(
                        f"Σειρά {row_num}: {category} / {description or '--Επιλέξτε'}"
                    )

        # Αν δεν υπάρχει κανένα πρόβλημα → επέστρεψε None
        if not missing and not missing_vine:
            return None

        # Δόμησε summary και details συνδυάζοντας τα δύο προβλήματα
        summary_text=(
                "Βρέθηκαν εγγραφές χωρίς έγκυρη Κατηγορία ΟΣΔΕ / Περιγραφή ή \n" 
                "εγγραφές αμπελιού χωρίς συμπληρωμένη τη στήλη 'Αμπέλι >3 ετών.\n\n"
                "Για τις παρακάτω γραμμές δεν μπορεί να υπολογιστεί η Τυπική Απόδοση."
            )
        
        

        details_parts = []
        if missing:
            details_parts.append(
                "Εγγραφές χωρίς έγκυρη Κατηγορία ΟΣΔΕ / Περιγραφή:\n"
                + "\n".join(f"- {item}" for item in missing)
            )
        if missing_vine:
            details_parts.append(
                "Εγγραφές αμπελιού χωρίς τιμή στη στήλη 'Αμπέλι >3 ετών':\n"
                + "\n".join(f"- {item}" for item in missing_vine)
            )
        details_text = "\n\n".join(details_parts)

        return summary_text, details_text
    
  
    
    def _build_missing_check_data_from_db_rows(self, afm, initial_rows):
        return {
            'afm': afm,
            'rows': [
                {
                    'category_osde': r[3] or '',
                    'description': r[4] or '',
                    'vine_over_3': r[10] or ''   # ΝΕΟ — index 10 = vine_over_3
                }
                for r in initial_rows
            ]
        }
    def _on_current_page_changed(self, index):
        """Όταν αλλάζει σελίδα, αν μπήκαμε στην ta_page ή mel_page δείξε τυχόν
        pending warning. Επίσης αποτρέπει το auto-focus σε widgets της νέας σελίδας."""
        if self.ui.stackedWidget.widget(index) is self.ta_page:
            self._show_pending_ta_warning()
        elif self.ui.stackedWidget.widget(index) is self.mel_page:
            self._show_pending_mel_warning()

        # Deferred focus reset — το Qt μεταφέρει auto-focus στο πρώτο focusable
        # widget της νέας σελίδας ως deferred event, οπότε το setFocus() εκτελείται
        # στον επόμενο γύρο event loop ώστε να επικρατήσει.
        QTimer.singleShot(0, self.setFocus)


    def _show_pending_ta_warning(self):
        """Εμφάνιση warning μόνο όταν η ta_page είναι η ενεργή σελίδα."""
        if not self._pending_ta_warning_message:
            return

        if self.ui.stackedWidget.currentWidget() is not self.ta_page:
            return

        summary_text, details_text = self._pending_ta_warning_message
        self._pending_ta_warning_message = None

        dialog = MissingCategoriesDialog(summary_text, details_text, self.ta_page)
        dialog.exec_()

    def _show_pending_mel_warning(self):
        """Εμφάνιση warning μόνο όταν η mel_page είναι η ενεργή σελίδα."""
        if not self._pending_mel_warning_message:
            return

        if self.ui.stackedWidget.currentWidget() is not self.mel_page:
            return

        summary_text, details_text = self._pending_mel_warning_message
        self._pending_mel_warning_message = None

        dialog = MissingCategoriesDialog(summary_text, details_text, self.mel_page)
        dialog.exec_()


    

    # def _load_ta_categories(self):
    #     """Φόρτωση ta.xls ΜΟΝΟ για έλεγχο ύπαρξης"""
    
        
    #     categories = set()
        
    #     try:
    #         ta_path = resource_path("data/ta.xlsx")
    #         wb = openpyxl.load_workbook(ta_path, read_only=True, data_only=True)
    #         ws = wb.active
            
    #         for row in ws.iter_rows(min_row=2, values_only=True):
    #             if len(row) < 3:
    #                 continue
                
    #             category = str(row[1]).strip() if row[1] else ''
    #             description = str(row[2]).strip() if row[2] else ''
                
    #             if category and description:
    #                 key = (norm(category), norm(description))
    #                 categories.add(key)
            
    #         wb.close()
        
    #     except Exception as e:
    #         print(f"Error loading ta.xlsx: {e}")
        
    #     return categories

    