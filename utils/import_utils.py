"""Validation και parsing λογική για xlsx import."""
import io
import openpyxl
from utils.excel_loader import norm, in_norm_set, LOCK_AMPELI_NORM

_ALLOWED_CERTS = {"--Επιλέξτε", "Συμβατικά", "Βιολογικά", "Ολοκληρωμένη", "ΠΟΠ/ΠΓΕ"}
_ALLOWED_VINE_XLSX = {"--Επιλέξτε", "Ναι", "Όχι"}
_MAX_QUANTITY = 9999999.99
_MAX_INT_7 = 9999999
_MAX_TEXT_LEN = 200


def build_canon_dicts(ta_value_mapping):
    """Φτιάχνει reverse-lookup dicts από TA_VALUE_MAPPING για canonicalization/validation."""
    canonical_pair_by_norm = {}
    canonical_cat_by_norm = {}
    valid_cats_cache = set()
    valid_descs_cache = set()
    for (cat, desc) in ta_value_mapping.keys():
        nc = norm(cat)
        nd = norm(desc)
        canonical_pair_by_norm[(nc, nd)] = (cat, desc)
        if nc not in canonical_cat_by_norm:
            canonical_cat_by_norm[nc] = cat
        valid_cats_cache.add(nc)
        valid_descs_cache.add(nd)
    return canonical_pair_by_norm, canonical_cat_by_norm, valid_cats_cache, valid_descs_cache


def _canonicalize_entry_row(entry_row, canonical_pair_by_norm, canonical_cat_by_norm):
    """Αντικαθιστά cat/desc με κανονικά αν ταιριάζουν normalized. Mutates in place."""
    cat = (entry_row.get('category_osde') or '').strip()
    desc = (entry_row.get('description') or '').strip()
    if not cat:
        return
    nc = norm(cat)
    if desc:
        nd = norm(desc)
        canonical = canonical_pair_by_norm.get((nc, nd))
        if canonical is not None:
            entry_row['category_osde'] = canonical[0]
            entry_row['description'] = canonical[1]
            return
    canonical_cat = canonical_cat_by_norm.get(nc)
    if canonical_cat is not None:
        entry_row['category_osde'] = canonical_cat


def _validate_import_row(entry_row, valid_cats_cache, valid_descs_cache):
    """Raises ValueError στο πρώτο άκυρο πεδίο."""
    cat = (entry_row.get('category_osde') or '').strip()
    if cat and norm(cat) not in valid_cats_cache and len(cat) > _MAX_TEXT_LEN:
        raise ValueError("Μη έγκυρα δεδομένα")

    desc = (entry_row.get('description') or '').strip()
    if desc and norm(desc) not in valid_descs_cache and len(desc) > _MAX_TEXT_LEN:
        raise ValueError("Μη έγκυρα δεδομένα")

    qty = entry_row.get('quantity', '')
    qty_str = str(qty).strip() if qty is not None else ''
    if qty_str:
        qty_norm = qty_str.replace(',', '.')
        try:
            qty_f = float(qty_norm)
        except (ValueError, TypeError):
            raise ValueError("Μη έγκυρα δεδομένα")
        if qty_f < 0 or qty_f > _MAX_QUANTITY:
            raise ValueError("Μη έγκυρα δεδομένα")
        if '.' in qty_norm and len(qty_norm.split('.', 1)[1]) > 2:
            raise ValueError("Μη έγκυρα δεδομένα")

    for field in ('trees_over_4', 'trees_under_4'):
        val = entry_row.get(field, '')
        val_str = str(val).strip() if val is not None else ''
        if val_str:
            try:
                val_f = float(val_str.replace(',', '.'))
            except (ValueError, TypeError):
                raise ValueError("Μη έγκυρα δεδομένα")
            if val_f != int(val_f):
                raise ValueError("Μη έγκυρα δεδομένα")
            if int(val_f) < 0 or int(val_f) > _MAX_INT_7:
                raise ValueError("Μη έγκυρα δεδομένα")

    cert = (entry_row.get('certification') or '').strip()
    if cert not in _ALLOWED_CERTS:
        raise ValueError("Μη έγκυρα δεδομένα")

    if cat and in_norm_set(cat, LOCK_AMPELI_NORM):
        vine = (entry_row.get('vine_over_3') or '').strip()
        if vine and vine not in _ALLOWED_VINE_XLSX:
            raise ValueError("Μη έγκυρα δεδομένα")


def read_excel_file(file_bytes, canonical_pair_by_norm, canonical_cat_by_norm,
                    valid_cats_cache, valid_descs_cache):
    """Διαβάζει xlsx bytes, επιστρέφει (import_data, skipped_afms). Raises ValueError σε σφάλμα."""
    if file_bytes[:2] != b'PK':
        raise ValueError("Το αρχείο δεν είναι έγκυρο Excel (.xlsx).")

    wb = None
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        if "TA Αρχικής" not in wb.sheetnames:
            raise ValueError("Δεν βρέθηκε το φύλλο 'TA Αρχικής' στο αρχείο!")
        ws = wb["TA Αρχικής"]

        headers = [' '.join(str(c.value).split()) if c.value else '' for c in ws[1]]

        col_map = {}
        for i, header in enumerate(headers):
            h = header.lower()
            if 'αφμ' in h or h == 'afm':
                col_map['afm'] = i
            elif 'επώνυμο' in h or 'surname' in h:
                col_map['surname'] = i
            elif 'όνομα' in h or 'name' in h:
                col_map['name'] = i
            elif 'κατηγορία οσδε' in h or 'category' in h:
                col_map['category_osde'] = i
            elif 'περιγραφή' in h or 'description' in h:
                col_map['description'] = i
            elif 'έκταση' in h or 'αριθμός ζώων' in h or 'quantity' in h:
                col_map['quantity'] = i
            elif 'βιολογικά' in h or 'certification' in h:
                col_map['certification'] = i
            elif 'δένδρα>=4' in h or 'δένδρα >=4' in h or 'trees_over_4' in h:
                col_map['trees_over_4'] = i
            elif 'δένδρα<4' in h or 'δένδρα <4' in h or 'trees_under_4' in h:
                col_map['trees_under_4'] = i
            elif 'αμπέλι>3' in h or 'αμπέλι >3' in h or 'vine_over_3' in h:
                col_map['vine_over_3'] = i

        REQUIRED = {
            'afm': "ΑΦΜ", 'name': "Όνομα", 'surname': "Επώνυμο",
            'category_osde': "Κατηγορία ΟΣΔΕ", 'description': "Περιγραφή",
            'quantity': "Έκταση/Αριθμός ζώων", 'certification': "Βιολογικά",
            'trees_over_4': "Δένδρα >=4 ετών", 'trees_under_4': "Δένδρα <4 ετών",
            'vine_over_3': "Αμπέλι >3 ετών",
        }
        missing = [lbl for k, lbl in REQUIRED.items() if k not in col_map]
        if missing:
            raise ValueError(f"Λείπουν υποχρεωτικές στήλες: {', '.join(missing)}")

        def get_cell(row, key):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return ''
            v = row[idx]
            return '' if (v is None or v == '') else str(v).strip()

        afm_groups = {}
        skipped_afms = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            afm = get_cell(row, 'afm')
            if not afm or len(afm) != 9:
                skipped_afms.append(afm)
                continue
            if afm not in afm_groups:
                afm_groups[afm] = {
                    'afm': afm,
                    'name': get_cell(row, 'name'),
                    'surname': get_cell(row, 'surname'),
                    'region': '--Επιλέξτε',
                    'rows': [],
                }
            entry_row = {
                'category_osde': get_cell(row, 'category_osde'),
                'description':   get_cell(row, 'description'),
                'quantity':      get_cell(row, 'quantity'),
                'certification': get_cell(row, 'certification'),
                'trees_over_4':  get_cell(row, 'trees_over_4'),
                'trees_under_4': get_cell(row, 'trees_under_4'),
                'vine_over_3':   get_cell(row, 'vine_over_3'),
                'typical_output': '', 'output_per_choice': '',
                'total_output': '', 'ta_productive': '',
                'ta_plant': '', 'ta_animal': '', 'ta_bees': '',
            }
            _canonicalize_entry_row(entry_row, canonical_pair_by_norm, canonical_cat_by_norm)
            _validate_import_row(entry_row, valid_cats_cache, valid_descs_cache)
            afm_groups[afm]['rows'].append(entry_row)

        return list(afm_groups.values()), skipped_afms
    finally:
        if wb is not None:
            wb.close()
