import sys
import os
import unicodedata
from openpyxl import load_workbook


def load_excel_data(filepath):
    """Φορτωση δεδομένων από το excel ta"""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    crop_to_variants = {}
    value_mapping = {}

    def _safe_round(v):
        """round(float(v), 2) ή None αν η τιμή δεν είναι αριθμητική."""
        if v is None:
            return None
        try:
            return round(float(v), 2)
        except (ValueError, TypeError):
            return None

    for row in ws.iter_rows(min_row=2, values_only=True):
        crop = row[1]
        variaty = row[2]
        value_4 = _safe_round(row[3])
        value_5 = _safe_round(row[4])

        if not crop or not variaty:
            continue

        crop = str(crop).strip()
        variaty = str(variaty).strip()
 
        crop_to_variants.setdefault(crop, set()).add(variaty)

        value_mapping[(crop, variaty)] = {
            "default": value_4,
            "aegean": value_5
        }

    return (
        {k: sorted(v) for k, v in crop_to_variants.items()},
        value_mapping
    )


def resource_path(relative_path):
    """
    Επιστρέφει σωστό path είτε τρέχεις .py είτε .exe (PyInstaller)
    """
    try:
        base_path = sys._MEIPASS  # PyInstaller
    except AttributeError:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

#set για συγκρίσεις στους πίνακες τυπικών αποδόσεων

FMZ_ZWIKI={
    "ΑΙΓΟΠΡΟΒΑΤΑ", 
    "ΒΟΟΕΙΔΗ", 
    "ΓΟΥΝΟΦΟΡΑ", 
    "ΘΗΡΑΜΑΤΑ", 
    "ΘΗΡΑΜΑΤΙΚΑ ΠΤΗΝΑ", 
    "ΙΠΠΟΕΙΔΗ","ΚΟΝΙΚΛΟΕΙΔΗ", 
    "ΟΡΝΙΘΟΕΙΔΗ",
    "ΧΟΙΡΟΙ",
    "ΧΩΡΟΙ ΕΚΤΡΟΦΗΣ ΣΑΛΙΓΚΑΡΙΩΝ"
}

FMZ_MELISSES={
    "ΚΥΨΕΛΕΣ ΜΕΛΙΣΣΩΝ - ΜΕΛΛΙΣΟΣΜΗΝΗ",
    "ΜΕΤΑΞΟΣΚΩΛΗΚΕΣ",
    
}

LOCK_AMPELI={
    "ΑΜΠΕΛΩΝΕΣ ΓΙΑ ΕΠΙΤΡΑΠΕΖΙΑ ΧΡΗΣΗ", 
    "ΑΜΠΕΛΩΝΕΣ ΓΙΑ ΠΑΡΑΓΩΓΗ ΟΙΝΟΥ", 
    "ΑΜΠΕΛΩΝΕΣ ΓΙΑ ΠΑΡΑΓΩΓΗ ΣΤΑΦΙΔΑΣ"
}

LOCK_TREES={
    "ΑΙΓΟΠΡΟΒΑΤΑ", 
    "ΒΟΟΕΙΔΗ", 
    "ΓΟΥΝΟΦΟΡΑ", 
    "ΘΗΡΑΜΑΤΑ", 
    "ΘΗΡΑΜΑΤΙΚΑ ΠΤΗΝΑ", 
    "ΙΠΠΟΕΙΔΗ","ΚΟΝΙΚΛΟΕΙΔΗ", 
    "ΟΡΝΙΘΟΕΙΔΗ",
    "ΧΟΙΡΟΙ",
    "ΧΩΡΟΙ ΕΚΤΡΟΦΗΣ ΣΑΛΙΓΚΑΡΙΩΝ",
    "ΑΜΠΕΛΩΝΕΣ ΓΙΑ ΕΠΙΤΡΑΠΕΖΙΑ ΧΡΗΣΗ", 
    "ΑΜΠΕΛΩΝΕΣ ΓΙΑ ΠΑΡΑΓΩΓΗ ΟΙΝΟΥ", 
    "ΑΜΠΕΛΩΝΕΣ ΓΙΑ ΠΑΡΑΓΩΓΗ ΣΤΑΦΙΔΑΣ",
    "--Επιλέξτε",
    "ΚΥΨΕΛΕΣ ΜΕΛΙΣΣΩΝ - ΜΕΛΛΙΣΟΣΜΗΝΗ",
    "ΜΕΤΑΞΟΣΚΩΛΗΚΕΣ",
    "ΙΧΘΥΟΚΑΛΛΙΕΡΓΕΙΑ",
    "ΑΓΡΑΝΑΠΑΥΣΗ",
    "ΑΡΑΒΟΣΙΤΟΣ",
    "ΑΡΑΒΟΣΙΤΟΣ ΕΝΣΙΡΩΣΗΣ",
    "ΒΑΜΒΑΚΙ",
    "ΒΙΟΜΗΧΑΝΙΚΗ ΚΑΝΝΑΒΗ",
    "ΒΙΟΜΗΧΑΝΙΚΗ ΚΑΝΝΑΒΗ ΕΚΤΟΣ ΚΟΙΝΟΤΙΚΟΥ ΚΑΤΑΛΟΓΟΥ",
    "ΒΟΣΚΟΤΟΠΟΙ",
    "ΓΕΩΜΗΛΑ",
    "ΓΗ ΠΟΥ ΔΕΝ ΕΝΤΑΣΣΕΤΑΙ ΣΕ ΚΑΛΛΙΕΡΓΗΤΙΚΗ ΔΡΑΣΤΗΡΙΟΤΗΤΑ",
    "ΕΚΤΑΣΕΙΣ ΜΕ ΑΠΕ",
    "ΕΛΑΙΟΥΧΟΙ ΣΠΟΡΟΙ",
    "ΕΝΕΡΓΕΙΑΚΕΣ ΚΑΛΛΙΕΡΓΕΙΕΣ",
    "ΕΞΑΙΡΕΤΙΚΗ ΠΕΡΙΣΤΑΣΗ",
    "ΚΑΠΝΟΣ",
    "ΖΑΧΑΡΟΤΕΥΤΛΑ",
    "ΚΗΠΕΥΤΙΚΑ",
    "ΚΗΠΕΥΤΙΚΑ ΥΠΟ ΚΑΛΥΨΗ",
    "ΚΤΗΝΟΤΡΟΦΙΚΑ ΦΥΤΑ ΓΙΑ ΖΩΟΤΡΟΦΕΣ",
    "ΛΙΝΟΣ ΚΛΩΣΤΙΚΟΣ",
    "ΛΙΝΟΣ ΜΗ ΚΛΩΣΤΙΚΟΣ",
    "ΛΟΙΠΑ ΣΙΤΗΡΑ",
    "ΜΗ ΕΠΙΛΕΞΙΜΕΣ ΕΚΤΑΣΕΙΣ",
    "ΜΗ ΕΠΙΛΕΞΙΜΕΣ ΕΚΤΑΣΕΙΣ ΑΠΟ ΑΝΑΔΑΣΩΤΕΕΣ ΕΚΤΑΣΕΙΣ ΚΥΡΩΜΕΝΩΝ",
    "ΜΗ ΕΠΙΛΕΞΙΜΕΣ ΕΚΤΑΣΕΙΣ ΑΠΟ ΚΥΡΩΜΕΝΟΥΣ ΔΑΣΙΚΟΥΣ",
    "ΜΗ ΚΑΛΛΙΕΡΓΗΣΙΜΕΣ ΕΚΤΑΣΕΙΣ ΛΟΓΩ ΦΥΣΙΚΩΝ ΚΑΤΑΣΤΡΟΦΩΝ",
    "ΟΣΠΡΙΑ ΒΡΩΣΙΜΑ",
    "ΠΑΡΕΚΚΛΙΣΗ ΑΓΡΑΝΑΠΑΥΣΗΣ",
    "ΡΥΖΙ",
    "ΣΚΛΗΡΟΣ ΣΙΤΟΣ",
    "ΣΠΑΡΑΓΓΙΑ",
    "ΤΟΜΑΤΑ ΒΙΟΜΗΧΑΝΙΚΗ",
    "ΦΑΡΜΑΚΕΥΤΙΚΗ ΚΑΝΝΑΒΗ"

}

PARAGWGIKA ={
    "ΕΡΙΦΙΑ ΘΗΛΥΚΑ",
    "ΑΙΓΕΣ",
    "ΑΜΝΟΙ ΘΗΛΥΚΟΙ",
    "ΠΡΟΒΑΤΙΝΕΣ",
    
}

PARAGWGIKA_CAT={"ΑΙΓΟΠΡΟΒΑΤΑ"}

LATIN_TO_GREEK = {
    "a": "α", "b": "β", "e": "ε", "h": "η", "i": "ι",
    "k": "κ", "m": "μ", "n": "ν", "o": "ο", "p": "ρ",
    "t": "τ", "y": "υ", "x": "χ", "z":"ζ", "u":"υ", "s":"σ"
}

# set και λίστες για το combobox της περιφέρειας
PERIFERIES=["--Επιλέξτε", 
            "Αν. Μακεδ. & Θράκη", 
            "Αττική", 
            "Βόρειο Αιγαίο", 
            "Δυτική Ελλάδα",
            "Δ. Μακεδονία",
            "Ήπειρος", 
            "Θεσσαλία",
            "Ιόνια Νησιά", 
            "Κ. Μακεδονία", 
            "Κρήτη", 
            "Νότιο Αιγαίο", 
            "Πελοπόννησος",
            "Στερεά Ελλάδα"

]

DEFAULT="--Επιλέξτε"


AEGEAN_PERIFERIES ={
    "Κρήτη", 
    "Νότιο Αιγαίο",
    "Βόρειο Αιγαίο"

}

ISLANDS={
    "Νότιο Αιγαίο",
    "Βόρειο Αιγαίο",
    "Ιόνια Νησιά"

}
#συναρτήσεις για normalize
def norm(text: str) -> str:
    """def για normalize κειμένου"""
    if not text:
        return ""

    text = text.strip().lower()

    # αφαίρεση τόνων
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")

    # αντικατάσταση λατινικών που μοιάζουν με ελληνικά
    text = "".join(LATIN_TO_GREEK.get(c, c) for c in text)

    # τελικό σ
    text = text.replace("ς", "σ")

    # καθάρισμα κενών
    text = " ".join(text.split())

    return text

def in_norm_set(value, norm_set):
    return norm(value) in norm_set

def contains_norm_keyword(value, norm_set):
    value_norm = norm(value)
    return any(k in value_norm for k in norm_set)

#normalize των set
FMZ_ZWIKI_NORM = {norm(x) for x in FMZ_ZWIKI}
FMZ_MELISSES_NORM = {norm(x) for x in FMZ_MELISSES}
LOCK_AMPELI_NORM = {norm(x) for x in LOCK_AMPELI}
LOCK_TREES_NORM = {norm(x) for x in LOCK_TREES}
PARAGWGIKA_NORM = {norm(x) for x in PARAGWGIKA}
PARAGWGIKA_CAT_NORM = {norm(x) for x in PARAGWGIKA_CAT}

NORM_YES=norm('ΝΑΙ')


